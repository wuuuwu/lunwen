from __future__ import annotations

import csv
import os
import re
import tempfile
import unicodedata
from collections.abc import Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.comments import Comment  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.table import (  # type: ignore[import-untyped]
    Table,
    TableStyleInfo,
)

from paper_reviewer.domain.batch import BatchItem, BatchRecord
from paper_reviewer.domain.rubric import RubricDimension
from paper_reviewer.domain.submission import SubmissionMetadata

_INVALID_WINDOWS_FILENAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_WHITESPACE = re.compile(r"\s+")
_WINDOWS_RESERVED = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{number}" for number in range(1, 10)),
    *(f"LPT{number}" for number in range(1, 10)),
}
_REPORT_SUFFIX = "_课程论文评测报告.pdf"
_PENDING_REVIEW_MARKER = "__待核对__"
_CRITICAL_FILENAME_FIELDS = frozenset({"student_name", "student_id", "paper_title"})
_METADATA_FIELD_LABELS = {
    "student_name": "姓名",
    "student_id": "学号",
    "paper_title": "题目",
}
_MAX_FILENAME_UTF16_UNITS = 240
_MAX_WINDOWS_PATH_UTF16_UNITS = 259
BATCH_SUMMARY_FILENAME = "课程论文评测汇总.csv"
BATCH_WORKBOOK_FILENAME = "课程论文评测汇总.xlsx"
BATCH_OUTPUT_OWNED_MESSAGE = "该报告输出目录已属于另一个课程论文批次；请选择新的空目录。"
BATCH_OUTPUT_SUMMARY_EXISTS_MESSAGE = (
    "该报告输出目录已存在课程论文评测汇总表，但没有可验证的批次归属；请选择新的空目录。"
)
BATCH_OUTPUT_OWNERSHIP_UNVERIFIABLE_MESSAGE = (
    "该报告输出目录的批次归属标记无法验证；请选择新的空目录。"
)


class BatchOutputConflictError(FileExistsError):
    """A safe, path-free conflict that prevents reusing a batch output folder."""


class BatchOutputOwnedByAnotherBatchError(BatchOutputConflictError):
    """The fixed summary path is already claimed by another batch."""

    def __init__(self) -> None:
        super().__init__(BATCH_OUTPUT_OWNED_MESSAGE)


class BatchOutputSummaryExistsError(BatchOutputConflictError):
    """A fixed summary exists without a matching ownership marker."""

    def __init__(self) -> None:
        super().__init__(BATCH_OUTPUT_SUMMARY_EXISTS_MESSAGE)


class BatchOutputOwnershipUnverifiableError(BatchOutputConflictError):
    """An ownership marker exists but cannot be read safely."""

    def __init__(self) -> None:
        super().__init__(BATCH_OUTPUT_OWNERSHIP_UNVERIFIABLE_MESSAGE)


def batch_output_conflict_message(
    output_dir: Path,
    *,
    batch_id: str | None = None,
) -> str | None:
    """Inspect the fixed batch output paths without creating or changing files.

    Omitting ``batch_id`` is appropriate when creating a new batch: any owner
    marker means that the directory is already in use.  Supplying the existing
    batch ID lets resume paths distinguish their own summary from a conflict.
    Only static, path-free messages are returned.
    """

    root = output_dir.resolve(strict=False)
    destination = root / BATCH_SUMMARY_FILENAME
    workbook = root / BATCH_WORKBOOK_FILENAME
    owner_path = _owner_path(destination)
    if owner_path.exists() or owner_path.is_symlink():
        if batch_id is None:
            return BATCH_OUTPUT_OWNED_MESSAGE
        try:
            owner = _read_output_owner(owner_path)
        except BatchOutputOwnershipUnverifiableError:
            return BATCH_OUTPUT_OWNERSHIP_UNVERIFIABLE_MESSAGE
        if owner != batch_id:
            return BATCH_OUTPUT_OWNED_MESSAGE
        return None
    if (
        destination.exists()
        or destination.is_symlink()
        or workbook.exists()
        or workbook.is_symlink()
    ):
        return BATCH_OUTPUT_SUMMARY_EXISTS_MESSAGE
    return None


def claim_batch_output_directory(output_dir: Path, batch_id: str) -> None:
    """Atomically claim the fixed summary path for ``batch_id``."""

    root = output_dir.resolve(strict=False)
    destination = root / BATCH_SUMMARY_FILENAME
    workbook = root / BATCH_WORKBOOK_FILENAME
    owner_path = _owner_path(destination)
    owner_existed = owner_path.exists() or owner_path.is_symlink()
    _claim_output_directory(destination, batch_id)
    if not owner_existed and (workbook.exists() or workbook.is_symlink()):
        owner_path.unlink(missing_ok=True)
        raise BatchOutputSummaryExistsError


def release_batch_output_directory_claim(output_dir: Path, batch_id: str) -> None:
    """Release a claim only when it still belongs to ``batch_id``."""

    owner_path = _owner_path(output_dir.resolve(strict=False) / BATCH_SUMMARY_FILENAME)
    try:
        owner = _read_output_owner(owner_path)
    except (FileNotFoundError, BatchOutputOwnershipUnverifiableError):
        return
    if owner != batch_id:
        return
    try:
        owner_path.unlink()
    except FileNotFoundError:
        return


def sanitize_filename_component(
    value: str,
    *,
    fallback: str,
    max_utf16_units: int = 60,
) -> str:
    normalized = unicodedata.normalize("NFKC", value)
    normalized = _WHITESPACE.sub(" ", normalized).strip()
    normalized = _INVALID_WINDOWS_FILENAME.sub("_", normalized).rstrip(" .")
    if not normalized:
        normalized = fallback
    if normalized.split(".", 1)[0].upper() in _WINDOWS_RESERVED:
        normalized = f"_{normalized}"
    normalized = _truncate_utf16(normalized, max_utf16_units).rstrip(" .")
    return normalized or fallback


def build_report_filename(
    metadata: SubmissionMetadata,
    run_id: str,
    *,
    source_filename: str | None = None,
) -> str:
    if _critical_metadata_needs_review(metadata):
        run_token = sanitize_filename_component(
            run_id[:8],
            fallback="run",
            max_utf16_units=8,
        )
        return f"待核对论文{_PENDING_REVIEW_MARKER}{run_token}{_REPORT_SUFFIX}"
    components = (
        sanitize_filename_component(metadata.student_name, fallback="未识别姓名"),
        sanitize_filename_component(metadata.student_id, fallback="未识别学号"),
        sanitize_filename_component(
            metadata.paper_title,
            fallback="未识别题目",
            max_utf16_units=90,
        ),
    )
    base = "_".join(components)
    available = _MAX_FILENAME_UTF16_UNITS - _utf16_units(_REPORT_SUFFIX)
    base = _truncate_utf16(base, available).rstrip(" ._") or "未识别论文"
    return f"{base}{_REPORT_SUFFIX}"


def allocate_report_path(
    output_dir: Path,
    metadata: SubmissionMetadata,
    run_id: str,
    *,
    source_filename: str | None = None,
) -> Path:
    """Choose a deterministic, non-overwriting report path."""

    output_dir = output_dir.resolve(strict=False)
    initial = build_report_filename(
        metadata,
        run_id,
        source_filename=source_filename,
    )
    filename = _fit_report_filename(
        output_dir,
        initial.removesuffix(_REPORT_SUFFIX),
        unique_suffix="",
        fallback="未识别论文",
    )
    existing = (
        {entry.name.casefold() for entry in output_dir.iterdir()}
        if output_dir.is_dir()
        else set()
    )
    if filename.casefold() not in existing:
        return output_dir / filename

    run_suffix = f"__{sanitize_filename_component(run_id[:8], fallback='run', max_utf16_units=8)}"
    suffix_units = _utf16_units(run_suffix + _REPORT_SUFFIX)
    stem = filename.removesuffix(_REPORT_SUFFIX)
    stem = _truncate_utf16(stem, _MAX_FILENAME_UTF16_UNITS - suffix_units).rstrip(" ._")
    candidate = _fit_report_filename(
        output_dir,
        stem,
        unique_suffix=run_suffix,
        fallback="报告",
    )
    counter = 2
    while candidate.casefold() in existing:
        numbered_suffix = f"{run_suffix}_{counter}"
        suffix_units = _utf16_units(numbered_suffix + _REPORT_SUFFIX)
        shortened = _truncate_utf16(stem, _MAX_FILENAME_UTF16_UNITS - suffix_units).rstrip(" ._")
        candidate = _fit_report_filename(
            output_dir,
            shortened,
            unique_suffix=numbered_suffix,
            fallback="报告",
        )
        counter += 1
    return output_dir / candidate


def is_allocated_report_filename(
    output_dir: Path,
    filename: str,
    metadata: SubmissionMetadata,
    run_id: str,
    *,
    source_filename: str | None = None,
) -> bool:
    """Recognize only names this allocator can produce for one exact run."""

    output_dir = output_dir.resolve(strict=False)
    initial = build_report_filename(
        metadata,
        run_id,
        source_filename=source_filename,
    )
    fitted_initial = _fit_report_filename(
        output_dir,
        initial.removesuffix(_REPORT_SUFFIX),
        unique_suffix="",
        fallback="未识别论文",
    )
    if filename.casefold() == fitted_initial.casefold():
        return True
    candidate_stem = filename.removesuffix(_REPORT_SUFFIX)
    if candidate_stem == filename:
        return False
    run_token = sanitize_filename_component(
        run_id[:8],
        fallback="run",
        max_utf16_units=8,
    )
    match = re.search(
        rf"__{re.escape(run_token)}(?:_(?P<number>[0-9]+))?$",
        candidate_stem,
        re.IGNORECASE,
    )
    if match is None:
        return False
    number_text = match.group("number")
    if number_text is not None and (number_text.startswith("0") or int(number_text) < 2):
        return False
    unique_suffix = f"__{run_token}"
    if number_text is not None:
        unique_suffix = f"{unique_suffix}_{number_text}"
    base_stem = fitted_initial.removesuffix(_REPORT_SUFFIX)
    suffix_units = _utf16_units(unique_suffix + _REPORT_SUFFIX)
    shortened = _truncate_utf16(
        base_stem,
        _MAX_FILENAME_UTF16_UNITS - suffix_units,
    ).rstrip(" ._")
    expected = _fit_report_filename(
        output_dir,
        shortened,
        unique_suffix=unique_suffix,
        fallback="报告",
    )
    return filename.casefold() == expected.casefold()


def write_batch_summary_csv(
    destination: Path,
    batch: BatchRecord,
    dimensions: Sequence[tuple[str, str]],
) -> None:
    """Atomically write a spreadsheet-safe, dynamically-columned batch summary."""

    destination.parent.mkdir(parents=True, exist_ok=True)
    _claim_output_directory(destination, batch.batch_id)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=destination.parent,
        prefix=f".{destination.name}.",
        suffix=".tmp",
    )
    temporary = Path(temporary_name)
    dimension_columns = _dimension_columns(dimensions)
    headers = [
        "原文件名",
        "姓名",
        "学号",
        "题目",
        "元数据置信度",
        "元数据待核对",
        "待核对字段",
        "人工已核对",
        "重复PDF内容",
        *(column for _, column in dimension_columns),
        "总分",
        "等级",
        "结论",
        "任务状态",
        "PDF文件名",
        "错误摘要",
    ]
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            for item in batch.items:
                writer.writerow(_csv_row(item, dimension_columns))
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)


def write_batch_summary_xlsx(
    destination: Path,
    batch: BatchRecord,
    dimensions: Sequence[RubricDimension],
) -> None:
    """Atomically write a styled, formula-safe Excel grade workbook."""

    destination.parent.mkdir(parents=True, exist_ok=True)
    claim_batch_output_directory(destination.parent, batch.batch_id)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=destination.parent,
        prefix=f".{destination.name}.",
        suffix=".tmp.xlsx",
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    workbook: Any | None = None
    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "成绩汇总"
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"

        dimension_columns = _dimension_columns(
            [(dimension.dimension_id, dimension.title) for dimension in dimensions]
        )
        headers = _summary_headers(dimension_columns)
        for column, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=column)
            _set_literal_text(cell, header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        dimension_by_id = {dimension.dimension_id: dimension for dimension in dimensions}
        dimension_header_start = (
            headers.index("重复PDF内容") + 2 if dimension_columns else 1
        )
        for offset, (dimension_id, _header) in enumerate(dimension_columns):
            dimension = dimension_by_id[dimension_id]
            cell = worksheet.cell(row=1, column=dimension_header_start + offset)
            cell.comment = Comment(
                "\n".join(
                    (
                        f"维度 ID：{dimension.dimension_id}",
                        f"权重：{_compact_number(dimension.weight)}%",
                        "分值范围："
                        f"{_compact_number(dimension.minimum_score)}–"
                        f"{_compact_number(dimension.maximum_score)}",
                    )
                ),
                "课程论文评测",
            )

        for row_number, item in enumerate(batch.items, start=2):
            values = _workbook_row(item, dimension_columns)
            for column, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row_number, column=column)
                value = values[header]
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.value = float(value)
                    cell.number_format = "0.00"
                else:
                    _set_literal_text(cell, "" if value is None else str(value))
                cell.alignment = Alignment(vertical="top", wrap_text=header in _WRAPPED_HEADERS)

        last_column = get_column_letter(len(headers))
        table_reference = f"A1:{last_column}{worksheet.max_row}"
        worksheet.auto_filter.ref = table_reference
        table = Table(displayName="CoursePaperScores", ref=table_reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        worksheet.row_dimensions[1].height = 34
        _fit_workbook_columns(worksheet, headers)
        worksheet.print_title_rows = "1:1"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        workbook.properties.title = "课程论文评测成绩汇总"
        workbook.save(temporary)
        with temporary.open("r+b") as handle:
            os.fsync(handle.fileno())
        os.replace(temporary, destination)
    finally:
        if workbook is not None:
            workbook.close()
        temporary.unlink(missing_ok=True)


_WRAPPED_HEADERS = frozenset({"原文件名", "题目", "待核对字段", "结论", "错误摘要"})


def _summary_headers(dimension_columns: Sequence[tuple[str, str]]) -> list[str]:
    return [
        "原文件名",
        "姓名",
        "学号",
        "题目",
        "元数据置信度",
        "元数据待核对",
        "待核对字段",
        "人工已核对",
        "重复PDF内容",
        *(column for _, column in dimension_columns),
        "总分",
        "等级",
        "结论",
        "任务状态",
        "PDF文件名",
        "错误摘要",
    ]


def _workbook_row(
    item: BatchItem,
    dimensions: Sequence[tuple[str, str]],
) -> dict[str, str | float | None]:
    metadata = item.metadata
    row: dict[str, str | float | None] = {
        "原文件名": item.source.filename,
        "姓名": metadata.student_name if metadata else "未识别姓名",
        "学号": metadata.student_id if metadata else "未识别学号",
        "题目": metadata.paper_title if metadata else "未识别题目",
        "元数据置信度": _metadata_confidence_number(metadata),
        "元数据待核对": "是" if metadata is None or metadata.needs_review else "否",
        "待核对字段": "、".join(_pending_metadata_labels(metadata)),
        "人工已核对": "是" if metadata is not None and metadata.human_reviewed else "否",
        "重复PDF内容": "是" if item.source.duplicate_sha256 else "否",
        "总分": item.total_score,
        "等级": item.grade or "",
        "结论": item.conclusion or "",
        "任务状态": item.status.value,
        "PDF文件名": item.report_path.name if item.report_path else "",
        "错误摘要": item.error or "",
    }
    for dimension_id, title in dimensions:
        row[title] = item.dimension_scores.get(dimension_id)
    return row


def _set_literal_text(cell: Any, value: str) -> None:
    cell.value = value
    cell.data_type = "s"
    cell.number_format = "@"
    cell.hyperlink = None


def _fit_workbook_columns(worksheet: Any, headers: Sequence[str]) -> None:
    preferred_minimums = {
        "原文件名": 20,
        "姓名": 10,
        "学号": 16,
        "题目": 28,
        "待核对字段": 16,
        "结论": 28,
        "PDF文件名": 24,
        "错误摘要": 28,
    }
    for column, header in enumerate(headers, start=1):
        width = preferred_minimums.get(header, 12)
        for row in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row, column=column).value
            if value is not None:
                width = max(width, _display_width(str(value)) + 2)
        worksheet.column_dimensions[get_column_letter(column)].width = min(width, 48)


def _display_width(value: str) -> int:
    return sum(
        2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
        for character in value
    )


def _compact_number(value: float) -> str:
    return f"{value:g}"


def _csv_row(item: BatchItem, dimensions: Sequence[tuple[str, str]]) -> dict[str, object]:
    metadata = item.metadata
    row: dict[str, object] = {
        "原文件名": _spreadsheet_safe(item.source.filename),
        "姓名": _spreadsheet_safe(metadata.student_name if metadata else "未识别姓名"),
        "学号": _spreadsheet_safe(metadata.student_id if metadata else "未识别学号"),
        "题目": _spreadsheet_safe(metadata.paper_title if metadata else "未识别题目"),
        "元数据置信度": _metadata_confidence(metadata),
        "元数据待核对": "是" if metadata is None or metadata.needs_review else "否",
        "待核对字段": "、".join(_pending_metadata_labels(metadata)),
        "人工已核对": "是" if metadata is not None and metadata.human_reviewed else "否",
        "重复PDF内容": "是" if item.source.duplicate_sha256 else "否",
        "总分": "" if item.total_score is None else item.total_score,
        "等级": _spreadsheet_safe(item.grade or ""),
        "结论": _spreadsheet_safe(item.conclusion or ""),
        "任务状态": item.status.value,
        "PDF文件名": _spreadsheet_safe(item.report_path.name if item.report_path else ""),
        "错误摘要": _spreadsheet_safe(item.error or ""),
    }
    for dimension_id, title in dimensions:
        value = item.dimension_scores.get(dimension_id)
        row[title] = "" if value is None else value
    return row


def _metadata_confidence(metadata: SubmissionMetadata | None) -> str:
    if metadata is None:
        return ""
    values = [detail.confidence for detail in metadata.field_evidence.values()]
    return "" if not values else f"{sum(values) / len(values):.2f}"


def _metadata_confidence_number(metadata: SubmissionMetadata | None) -> float | None:
    if metadata is None:
        return None
    values = [detail.confidence for detail in metadata.field_evidence.values()]
    return None if not values else sum(values) / len(values)


def _pending_metadata_labels(metadata: SubmissionMetadata | None) -> list[str]:
    fields = (
        _METADATA_FIELD_LABELS
        if metadata is None
        else metadata.pending_review_fields
    )
    return [_METADATA_FIELD_LABELS[field] for field in fields]


def _critical_metadata_needs_review(metadata: SubmissionMetadata) -> bool:
    return bool(_CRITICAL_FILENAME_FIELDS.intersection(metadata.pending_review_fields))


def _dimension_columns(dimensions: Sequence[tuple[str, str]]) -> list[tuple[str, str]]:
    title_counts: dict[str, int] = {}
    for _, title in dimensions:
        title_counts[title] = title_counts.get(title, 0) + 1
    return [
        (
            dimension_id,
            _spreadsheet_safe(
                title if title_counts[title] == 1 else f"{title}（{dimension_id}）"
            ),
        )
        for dimension_id, title in dimensions
    ]


def _claim_output_directory(destination: Path, batch_id: str) -> None:
    """Claim one summary path without monopolizing the entire output folder."""

    owner_path = _owner_path(destination)
    try:
        descriptor = os.open(
            owner_path,
            os.O_CREAT | os.O_EXCL | os.O_WRONLY,
            0o600,
        )
    except FileExistsError:
        try:
            owner = _read_output_owner(owner_path)
        except BatchOutputOwnershipUnverifiableError:
            raise
        if owner != batch_id:
            raise BatchOutputOwnedByAnotherBatchError from None
        return
    try:
        with os.fdopen(descriptor, "w", encoding="ascii", newline="") as handle:
            handle.write(batch_id)
            handle.flush()
            os.fsync(handle.fileno())
    except Exception:
        owner_path.unlink(missing_ok=True)
        raise
    if destination.exists() or destination.is_symlink():
        owner_path.unlink(missing_ok=True)
        raise BatchOutputSummaryExistsError


def _owner_path(destination: Path) -> Path:
    return destination.with_name(f".{destination.name}.owner")


def _read_output_owner(owner_path: Path) -> str:
    if owner_path.is_symlink():
        raise BatchOutputOwnershipUnverifiableError
    try:
        with owner_path.open("r", encoding="ascii", newline="") as handle:
            owner = handle.read(129)
            if handle.read(1):
                raise BatchOutputOwnershipUnverifiableError
    except FileNotFoundError:
        raise
    except (OSError, UnicodeError) as error:
        raise BatchOutputOwnershipUnverifiableError from error
    owner = owner.strip()
    if not owner or len(owner) > 128 or any(character.isspace() for character in owner):
        raise BatchOutputOwnershipUnverifiableError
    return owner


def _spreadsheet_safe(value: str) -> str:
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _truncate_utf16(value: str, maximum_units: int) -> str:
    if maximum_units <= 0:
        return ""
    result: list[str] = []
    used = 0
    for character in value:
        units = _utf16_units(character)
        if used + units > maximum_units:
            break
        result.append(character)
        used += units
    return "".join(result)


def _utf16_units(value: str) -> int:
    return len(value.encode("utf-16-le")) // 2


def _fit_report_filename(
    output_dir: Path,
    stem: str,
    *,
    unique_suffix: str,
    fallback: str,
) -> str:
    directory_units = _utf16_units(str(output_dir)) + 1
    available = min(_MAX_FILENAME_UTF16_UNITS, _MAX_WINDOWS_PATH_UTF16_UNITS - directory_units)
    tail = unique_suffix + _REPORT_SUFFIX
    if available <= _utf16_units(tail):
        raise OSError("batch output directory path is too long")
    fitted = _truncate_utf16(stem, available - _utf16_units(tail)).rstrip(" ._")
    if not fitted:
        fitted = _truncate_utf16(fallback, available - _utf16_units(tail))
    return f"{fitted}{tail}"
