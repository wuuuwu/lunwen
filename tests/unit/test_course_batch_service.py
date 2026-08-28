from __future__ import annotations

import asyncio
import hashlib
from pathlib import Path
from threading import Event
from types import SimpleNamespace
from typing import Any

import pymupdf
import pytest
from openpyxl import load_workbook

import paper_reviewer.application.service as service_module
from paper_reviewer.application.artifacts import RunArtifactStore
from paper_reviewer.application.batch_output import (
    BATCH_SUMMARY_FILENAME,
    BATCH_WORKBOOK_FILENAME,
    BatchOutputConflictError,
    BatchOutputOwnedByAnotherBatchError,
    BatchOutputSummaryExistsError,
    allocate_report_path,
    batch_output_conflict_message,
    build_report_filename,
    claim_batch_output_directory,
)
from paper_reviewer.application.batch_store import (
    BatchExecutionInProgressError,
    BatchStore,
    scan_batch_sources,
)
from paper_reviewer.application.metadata_recheck import submission_metadata_sha256
from paper_reviewer.application.models import MetadataRecheckDecision
from paper_reviewer.application.service import (
    BATCH_WORKBOOK_LOCKED_MESSAGE,
    BatchWorkbookExportError,
    ReviewApplicationService,
)
from paper_reviewer.config import Settings, load_review_profile, load_rubric
from paper_reviewer.domain.batch import (
    BatchItem,
    BatchItemStatus,
    BatchRecord,
    BatchReviewRequest,
    BatchStatus,
)
from paper_reviewer.domain.provider import (
    ModelApiProtocol,
    ProviderSnapshot,
    endpoint_fingerprint,
)
from paper_reviewer.domain.run import RunRecord, RunStatus
from paper_reviewer.domain.submission import (
    SUBMISSION_METADATA_FIELDS,
    SubmissionFieldEvidence,
    SubmissionMetadata,
    SubmissionMetadataSource,
)

PROJECT_ROOT = Path(__file__).resolve().parents[2]
COURSE_RUBRIC = PROJECT_ROOT / "configs" / "rubrics" / "course_paper_v1.yaml"
COURSE_PROFILE = PROJECT_ROOT / "configs" / "review_profiles" / "course_paper_reviewers_v1.yaml"


class MemoryBatchStore:
    """Small persistence double that keeps the real service transitions observable."""

    def __init__(self) -> None:
        self.record: BatchRecord | None = None
        self.saves = 0

    def create(self, record: BatchRecord) -> None:
        assert self.record is None
        self.record = record.model_copy(deep=True)

    def save(self, record: BatchRecord) -> None:
        self.saves += 1
        self.record = record.model_copy(deep=True)

    def load(self, batch_id: str) -> BatchRecord:
        assert self.record is not None
        assert self.record.batch_id == batch_id
        return self.record.model_copy(deep=True)

    def list_records(self) -> list[BatchRecord]:
        return [self.record.model_copy(deep=True)] if self.record is not None else []


class FakeProviders:
    def __init__(self, snapshot: ProviderSnapshot, *, api_key: str | None = "test-key") -> None:
        self.snapshot_value = snapshot
        self.api_key = api_key

    def snapshot(self, provider: str, model: str) -> ProviderSnapshot:
        assert provider == self.snapshot_value.provider_ref
        assert model == self.snapshot_value.model
        return self.snapshot_value

    def get_snapshot_api_key(self, snapshot: ProviderSnapshot) -> str | None:
        assert snapshot == self.snapshot_value
        return self.api_key


def _provider(model: str = "gpt-test") -> ProviderSnapshot:
    base_url = "https://api.openai.com/v1"
    return ProviderSnapshot(
        provider_ref="openai",
        display_name="OpenAI",
        protocol=ModelApiProtocol.CHAT_COMPLETIONS,
        base_url=base_url,
        endpoint_fingerprint=endpoint_fingerprint(base_url, ModelApiProtocol.CHAT_COMPLETIONS),
        model=model,
    )


def _request(tmp_path: Path, *, output_dir: Path | None = None) -> BatchReviewRequest:
    source_dir = tmp_path / "papers"
    source_dir.mkdir(exist_ok=True)
    return BatchReviewRequest(
        source_dir=source_dir,
        output_dir=output_dir or tmp_path / "reports",
        provider="openai",
        model="gpt-test",
        rubric=COURSE_RUBRIC,
        profile=COURSE_PROFILE,
        cloud_processing_authorized=True,
        pii_output_authorized=True,
        external_search=False,
    )


def _record(
    tmp_path: Path,
    statuses: list[BatchItemStatus],
    *,
    batch_status: BatchStatus = BatchStatus.PAUSED,
    current_item_id: str | None = None,
) -> BatchRecord:
    request = _request(tmp_path)
    for index in range(len(statuses)):
        (request.source_dir / f"paper-{index}.pdf").write_bytes(b"%PDF-1.4\n")
    items = scan_batch_sources(request)
    for item, status in zip(items, statuses, strict=True):
        item.status = status
        if status is BatchItemStatus.FAILED:
            item.error = "previous safe error"
    return BatchRecord(
        batch_id="batch-test",
        status=batch_status,
        request=request,
        rubric_snapshot=load_rubric(COURSE_RUBRIC),
        profile_snapshot=load_review_profile(COURSE_PROFILE),
        provider_snapshot=_provider(),
        items=items,
        current_item_id=current_item_id,
    )


def _service(
    tmp_path: Path,
    store: MemoryBatchStore | BatchStore,
    *,
    api_key: str | None = "test-key",
) -> ReviewApplicationService:
    service = ReviewApplicationService.__new__(ReviewApplicationService)
    service.paths = SimpleNamespace(batches_dir=tmp_path / "batches")
    service.settings = Settings(
        database_url=f"sqlite+aiosqlite:///{(tmp_path / 'review.db').as_posix()}",
        runs_dir=tmp_path / "runs",
    )
    service.providers = FakeProviders(_provider(), api_key=api_key)
    service._batch_store = lambda: store  # type: ignore[method-assign]
    return service


async def _no_output(*_args: Any, **_kwargs: Any) -> None:
    return None


def _metadata(name: str = "张三") -> SubmissionMetadata:
    return SubmissionMetadata(
        student_name=name,
        student_id="20260001",
        major="公共管理",
        paper_title="课程论文",
        field_evidence={
            field: SubmissionFieldEvidence(
                source=SubmissionMetadataSource.COVER_LABEL,
                confidence=0.95,
            )
            for field in SUBMISSION_METADATA_FIELDS
        },
    )


@pytest.mark.asyncio
async def test_create_batch_freezes_rubric_profile_provider_and_persists_manifest(
    tmp_path: Path,
) -> None:
    store = MemoryBatchStore()
    service = _service(tmp_path, store)
    (tmp_path / "papers").mkdir()
    (tmp_path / "papers" / "paper.pdf").write_bytes(b"%PDF-1.4\n")

    record = await service.create_batch(_request(tmp_path))

    assert record.status is BatchStatus.CREATED
    persisted = store.load(record.batch_id)
    assert persisted.rubric_snapshot == load_rubric(COURSE_RUBRIC)
    assert persisted.profile_snapshot == load_review_profile(COURSE_PROFILE)
    assert persisted.provider_snapshot == _provider()
    assert persisted.request.provider == "openai"
    assert persisted.provider_snapshot.model == "gpt-test"
    assert persisted.summary_path is not None
    assert persisted.summary_path.name == "课程论文评测汇总.csv"
    assert persisted.workbook_path is not None
    assert persisted.workbook_path.name == BATCH_WORKBOOK_FILENAME
    assert persisted.workbook_path.is_file()
    assert persisted.workbook_export_error is None


@pytest.mark.asyncio
async def test_create_batch_keeps_running_state_when_workbook_refresh_is_locked(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = MemoryBatchStore()
    service = _service(tmp_path, store)
    (tmp_path / "papers").mkdir()
    (tmp_path / "papers" / "paper.pdf").write_bytes(b"%PDF-1.4\n")

    def locked(*_args: Any, **_kwargs: Any) -> None:
        raise PermissionError("private locked path")

    monkeypatch.setattr(service_module, "write_batch_summary_xlsx", locked)

    record = await service.create_batch(_request(tmp_path))

    assert record.status is BatchStatus.CREATED
    assert record.error is None
    assert record.summary_path is not None and record.summary_path.is_file()
    assert record.workbook_export_error == BATCH_WORKBOOK_LOCKED_MESSAGE
    assert "path" not in record.workbook_export_error
    assert store.load(record.batch_id).workbook_export_error == BATCH_WORKBOOK_LOCKED_MESSAGE


@pytest.mark.asyncio
async def test_manual_workbook_export_supports_historical_batch_and_clears_warning(
    tmp_path: Path,
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.COMPLETED], batch_status=BatchStatus.COMPLETED)
    record.workbook_path = None
    record.workbook_export_error = "旧警告"
    record.items[0].metadata = _metadata()
    record.items[0].dimension_scores = {"task_completion": 81.5}
    record.items[0].total_score = 81.5
    store.create(record)
    service = _service(tmp_path, store)

    destination = await service.export_batch_workbook(record.batch_id)

    assert destination.name == BATCH_WORKBOOK_FILENAME
    assert destination.is_file()
    persisted = store.load(record.batch_id)
    assert persisted.status is BatchStatus.COMPLETED
    assert persisted.workbook_path == destination
    assert persisted.workbook_export_error is None
    workbook = load_workbook(destination, data_only=False)
    try:
        worksheet = workbook["成绩汇总"]
        headers = [cell.value for cell in worksheet[1]]
        assert worksheet.cell(row=2, column=headers.index("姓名") + 1).value == "张三"
        assert worksheet.cell(row=2, column=headers.index("总分") + 1).value == 81.5
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_manual_workbook_export_persists_safe_locked_warning_without_status_change(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = MemoryBatchStore()
    record = _record(
        tmp_path,
        [BatchItemStatus.COMPLETED],
        batch_status=BatchStatus.COMPLETED,
    )
    store.create(record)
    service = _service(tmp_path, store)

    def locked(*_args: Any, **_kwargs: Any) -> None:
        raise PermissionError("C:/private/student-list.xlsx")

    monkeypatch.setattr(service_module, "write_batch_summary_xlsx", locked)

    with pytest.raises(BatchWorkbookExportError) as captured:
        await service.export_batch_workbook(record.batch_id)

    assert str(captured.value) == BATCH_WORKBOOK_LOCKED_MESSAGE
    assert "private" not in str(captured.value)
    persisted = store.load(record.batch_id)
    assert persisted.status is BatchStatus.COMPLETED
    assert persisted.error is None
    assert persisted.workbook_export_error == BATCH_WORKBOOK_LOCKED_MESSAGE


@pytest.mark.asyncio
async def test_create_batch_rejects_unowned_existing_summary_before_persisting(
    tmp_path: Path,
) -> None:
    store = MemoryBatchStore()
    service = _service(tmp_path, store)
    request = _request(tmp_path)
    (request.source_dir / "paper.pdf").write_bytes(b"%PDF-1.4\n")
    request.output_dir.mkdir()
    (request.output_dir / BATCH_SUMMARY_FILENAME).write_bytes(b"previous summary")

    with pytest.raises(BatchOutputSummaryExistsError):
        await service.create_batch(request)

    assert store.record is None
    assert not service.paths.batches_dir.exists()


@pytest.mark.asyncio
async def test_create_batch_rejects_another_owner_before_persisting(
    tmp_path: Path,
) -> None:
    store = MemoryBatchStore()
    service = _service(tmp_path, store)
    request = _request(tmp_path)
    (request.source_dir / "paper.pdf").write_bytes(b"%PDF-1.4\n")
    request.output_dir.mkdir()
    claim_batch_output_directory(request.output_dir, "previous-batch")

    with pytest.raises(BatchOutputOwnedByAnotherBatchError):
        await service.create_batch(request)

    assert store.record is None
    assert not service.paths.batches_dir.exists()


@pytest.mark.asyncio
async def test_competing_batch_creators_do_not_both_persist_the_same_output(
    tmp_path: Path,
) -> None:
    request = _request(tmp_path)
    (request.source_dir / "paper.pdf").write_bytes(b"%PDF-1.4\n")
    first_store = MemoryBatchStore()
    second_store = MemoryBatchStore()
    first_service = _service(tmp_path / "first-state", first_store)
    second_service = _service(tmp_path / "second-state", second_store)

    results = await asyncio.gather(
        first_service.create_batch(request),
        second_service.create_batch(request),
        return_exceptions=True,
    )

    assert sum(isinstance(result, BatchRecord) for result in results) == 1
    conflict = next(result for result in results if isinstance(result, BaseException))
    assert isinstance(conflict, BatchOutputConflictError)
    assert sum(store.record is not None for store in (first_store, second_store)) == 1


@pytest.mark.asyncio
async def test_create_batch_releases_claim_without_masking_store_failure(
    tmp_path: Path,
) -> None:
    class FailingCreateStore(MemoryBatchStore):
        def create(self, record: BatchRecord) -> None:
            del record
            raise RuntimeError("original persistence failure")

    store = FailingCreateStore()
    service = _service(tmp_path, store)
    request = _request(tmp_path)
    (request.source_dir / "paper.pdf").write_bytes(b"%PDF-1.4\n")

    with pytest.raises(RuntimeError, match="original persistence failure"):
        await service.create_batch(request)

    assert store.record is None
    assert batch_output_conflict_message(request.output_dir) is None


@pytest.mark.asyncio
async def test_claim_cleanup_failure_does_not_mask_store_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class FailingCreateStore(MemoryBatchStore):
        def create(self, record: BatchRecord) -> None:
            del record
            raise RuntimeError("original persistence failure")

    store = FailingCreateStore()
    service = _service(tmp_path, store)
    request = _request(tmp_path)
    (request.source_dir / "paper.pdf").write_bytes(b"%PDF-1.4\n")

    def fail_cleanup(_output_dir: Path, _batch_id: str) -> None:
        raise PermissionError("private output owner path")

    monkeypatch.setattr(
        "paper_reviewer.application.service.release_batch_output_directory_claim",
        fail_cleanup,
    )

    with pytest.raises(RuntimeError, match="original persistence failure"):
        await service.create_batch(request)

    assert store.record is None


@pytest.mark.asyncio
async def test_cancelled_create_waits_for_manifest_before_preserving_claim(
    tmp_path: Path,
) -> None:
    started = Event()
    allow_finish = Event()

    class BlockingCreateStore(MemoryBatchStore):
        def create(self, record: BatchRecord) -> None:
            started.set()
            assert allow_finish.wait(timeout=5)
            super().create(record)

    store = BlockingCreateStore()
    service = _service(tmp_path, store)
    request = _request(tmp_path)
    (request.source_dir / "paper.pdf").write_bytes(b"%PDF-1.4\n")
    creation = asyncio.create_task(service.create_batch(request))
    assert await asyncio.to_thread(started.wait, 5)

    creation.cancel()
    await asyncio.sleep(0)
    assert not creation.done()
    assert batch_output_conflict_message(request.output_dir) is not None

    allow_finish.set()
    with pytest.raises(asyncio.CancelledError):
        await creation

    assert store.record is not None
    assert batch_output_conflict_message(request.output_dir) is not None


@pytest.mark.asyncio
async def test_cancelled_create_waits_for_claim_and_manifest_as_one_operation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    started = Event()
    allow_finish = Event()
    store = MemoryBatchStore()
    service = _service(tmp_path, store)
    request = _request(tmp_path)
    (request.source_dir / "paper.pdf").write_bytes(b"%PDF-1.4\n")

    def blocking_claim(output_dir: Path, batch_id: str) -> None:
        claim_batch_output_directory(output_dir, batch_id)
        started.set()
        assert allow_finish.wait(timeout=5)

    monkeypatch.setattr(
        "paper_reviewer.application.service.claim_batch_output_directory",
        blocking_claim,
    )
    creation = asyncio.create_task(service.create_batch(request))
    assert await asyncio.to_thread(started.wait, 5)

    creation.cancel()
    await asyncio.sleep(0)
    assert not creation.done()
    assert store.record is None
    assert batch_output_conflict_message(request.output_dir) is not None

    allow_finish.set()
    with pytest.raises(asyncio.CancelledError):
        await creation

    assert store.record is not None
    assert batch_output_conflict_message(request.output_dir) is not None


@pytest.mark.asyncio
async def test_create_and_run_batch_expose_initial_csv_failure_before_any_model_call(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    service = _service(tmp_path, store)
    (tmp_path / "papers").mkdir()
    (tmp_path / "papers" / "paper.pdf").write_bytes(b"%PDF-1.4\n")

    def fail_csv(_record: BatchRecord) -> None:
        raise PermissionError("output locked")

    monkeypatch.setattr(
        "paper_reviewer.application.service._write_batch_csv",
        fail_csv,
    )

    created = await service.create_batch(_request(tmp_path))

    assert created.status is BatchStatus.PAUSED
    assert created.error

    monkeypatch.setattr(
        service,
        "_run_batch_item",
        lambda *_args, **_kwargs: pytest.fail("CSV preflight must run before the model"),
    )
    resumed = await service.run_batch(created.batch_id)

    assert resumed.status is BatchStatus.PAUSED
    assert resumed.items[0].status is BatchItemStatus.QUEUED


@pytest.mark.asyncio
async def test_run_batch_is_sequential_and_single_item_failure_does_not_stop_later_items(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.QUEUED] * 3)
    store.create(record)
    service = _service(tmp_path, store)
    order: list[str] = []

    async def run_item(
        _record_arg: BatchRecord,
        item: BatchItem,
        *,
        store: MemoryBatchStore,
        event_sink: Any,
    ) -> None:
        del store, event_sink
        order.append(item.source.filename)
        if item.source.filename == "paper-1.pdf":
            raise RuntimeError("paper-specific parse failure")
        item.status = BatchItemStatus.COMPLETED

    monkeypatch.setattr(service, "_run_batch_item", run_item)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )
    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", lambda _record: None)

    result = await service.run_batch(record.batch_id)

    assert order == ["paper-0.pdf", "paper-1.pdf", "paper-2.pdf"]
    assert [item.status for item in result.items] == [
        BatchItemStatus.COMPLETED,
        BatchItemStatus.FAILED,
        BatchItemStatus.COMPLETED,
    ]
    assert result.status is BatchStatus.COMPLETED_WITH_ERRORS
    assert result.error is None


@pytest.mark.asyncio
async def test_run_batch_completes_when_excel_workbook_is_open(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.QUEUED])
    store.create(record)
    service = _service(tmp_path, store)

    async def complete_item(
        _record_arg: BatchRecord,
        item: BatchItem,
        **_kwargs: Any,
    ) -> None:
        item.status = BatchItemStatus.COMPLETED
        item.metadata = _metadata()
        item.dimension_scores = {"task_completion": 80.0}
        item.total_score = 80.0

    def workbook_locked(*_args: Any, **_kwargs: Any) -> None:
        raise PermissionError("spreadsheet is open")

    monkeypatch.setattr(service, "_run_batch_item", complete_item)
    monkeypatch.setattr(service_module, "write_batch_summary_xlsx", workbook_locked)
    monkeypatch.setattr(service_module, "validate_source_snapshot", lambda _source: None)

    result = await service.run_batch(record.batch_id)

    assert result.status is BatchStatus.COMPLETED
    assert result.error is None
    assert result.items[0].status is BatchItemStatus.COMPLETED
    assert result.summary_path is not None and result.summary_path.is_file()
    assert result.workbook_export_error == BATCH_WORKBOOK_LOCKED_MESSAGE
    assert store.load(record.batch_id).workbook_export_error == BATCH_WORKBOOK_LOCKED_MESSAGE


@pytest.mark.asyncio
async def test_run_batch_pauses_on_shared_provider_error_and_leaves_later_items_queued(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.QUEUED, BatchItemStatus.QUEUED])
    store.create(record)
    service = _service(tmp_path, store)

    class AuthenticationError(RuntimeError):
        status_code = 401

    async def run_item(*_args: Any, **_kwargs: Any) -> None:
        raise AuthenticationError("Bearer secret must never be persisted")

    monkeypatch.setattr(service, "_run_batch_item", run_item)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )
    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", lambda _record: None)

    result = await service.run_batch(record.batch_id)

    assert result.status is BatchStatus.PAUSED
    assert result.current_item_id == result.items[0].item_id
    assert result.items[0].status is BatchItemStatus.RUNNING
    assert result.items[0].error == "Provider 认证失败；请检查 API Key 和账号权限。"
    assert result.items[1].status is BatchItemStatus.QUEUED
    assert "Bearer" not in (result.error or "")


@pytest.mark.asyncio
async def test_pause_batch_marks_current_item_cancelled_and_preserves_checkpoint(
    tmp_path: Path,
) -> None:
    current_id = "not-used-until-overwritten"
    store = MemoryBatchStore()
    record = _record(
        tmp_path,
        [BatchItemStatus.RUNNING, BatchItemStatus.QUEUED],
        batch_status=BatchStatus.RUNNING,
    )
    current_id = record.items[0].item_id
    record.current_item_id = current_id
    store.create(record)
    service = _service(tmp_path, store)

    result = await service.pause_batch(record.batch_id)

    assert result.status is BatchStatus.PAUSED
    assert result.current_item_id == current_id
    assert result.items[0].status is BatchItemStatus.CANCELLED
    assert store.load(record.batch_id).items[0].status is BatchItemStatus.CANCELLED


@pytest.mark.asyncio
async def test_resume_batch_skips_completed_items(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.COMPLETED, BatchItemStatus.QUEUED])
    store.create(record)
    service = _service(tmp_path, store)
    called: list[str] = []

    async def run_item(_record_arg: BatchRecord, item: BatchItem, **_kwargs: Any) -> None:
        called.append(item.source.filename)
        item.status = BatchItemStatus.COMPLETED

    monkeypatch.setattr(service, "_run_batch_item", run_item)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )
    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", lambda _record: None)

    result = await service.resume_batch(record.batch_id)

    assert called == ["paper-1.pdf"]
    assert result.items[0].status is BatchItemStatus.COMPLETED
    assert result.items[1].status is BatchItemStatus.COMPLETED


@pytest.mark.asyncio
async def test_resume_revalidates_persisted_cloud_authorization(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.QUEUED])
    record.request.cloud_processing_authorized = False
    store.create(record)
    service = _service(tmp_path, store)
    monkeypatch.setattr(
        service,
        "_run_batch_item",
        lambda *_args, **_kwargs: pytest.fail("unauthorized batch must not start a run"),
    )

    result = await service.resume_batch(record.batch_id)

    assert result.status is BatchStatus.PAUSED
    assert result.items[0].status is BatchItemStatus.QUEUED


@pytest.mark.asyncio
async def test_retry_failed_items_does_not_retry_or_run_queued_items(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(
        tmp_path,
        [BatchItemStatus.COMPLETED, BatchItemStatus.FAILED, BatchItemStatus.QUEUED],
        batch_status=BatchStatus.COMPLETED_WITH_ERRORS,
    )
    store.create(record)
    service = _service(tmp_path, store)
    called: list[str] = []

    async def run_item(_record_arg: BatchRecord, item: BatchItem, **_kwargs: Any) -> None:
        called.append(item.source.filename)
        item.status = BatchItemStatus.COMPLETED

    monkeypatch.setattr(service, "_run_batch_item", run_item)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )
    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", lambda _record: None)

    result = await service.retry_failed_items(record.batch_id)

    assert called == ["paper-1.pdf"]
    assert result.items[1].status is BatchItemStatus.COMPLETED
    assert result.items[2].status is BatchItemStatus.QUEUED
    assert result.retry_item_ids is None
    assert result.status is BatchStatus.PAUSED


@pytest.mark.asyncio
async def test_retry_scope_survives_cancellation_and_resume_runs_only_selected_items(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(
        tmp_path,
        [BatchItemStatus.COMPLETED, BatchItemStatus.FAILED, BatchItemStatus.QUEUED],
        batch_status=BatchStatus.COMPLETED_WITH_ERRORS,
    )
    store.create(record)
    service = _service(tmp_path, store)

    async def cancel_item(*_args: Any, **_kwargs: Any) -> None:
        raise asyncio.CancelledError

    monkeypatch.setattr(service, "_run_batch_item", cancel_item)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )
    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", lambda _record: None)

    with pytest.raises(asyncio.CancelledError):
        await service.retry_failed_items(record.batch_id)

    persisted = store.load(record.batch_id)
    selected_id = persisted.items[1].item_id
    assert persisted.retry_item_ids == [selected_id]
    assert persisted.items[1].status is BatchItemStatus.CANCELLED
    assert persisted.items[2].status is BatchItemStatus.QUEUED

    called: list[str] = []

    async def complete_item(_record: BatchRecord, item: BatchItem, **_kwargs: Any) -> None:
        called.append(item.source.filename)
        item.status = BatchItemStatus.COMPLETED

    monkeypatch.setattr(service, "_run_batch_item", complete_item)
    resumed = await service.resume_batch(record.batch_id)

    assert called == ["paper-1.pdf"]
    assert resumed.retry_item_ids is None
    assert resumed.items[2].status is BatchItemStatus.QUEUED
    assert resumed.status is BatchStatus.PAUSED


@pytest.mark.asyncio
async def test_retry_scope_survives_shared_error_and_resume_stays_scoped(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(
        tmp_path,
        [BatchItemStatus.COMPLETED, BatchItemStatus.FAILED, BatchItemStatus.QUEUED],
        batch_status=BatchStatus.COMPLETED_WITH_ERRORS,
    )
    store.create(record)
    service = _service(tmp_path, store)

    class AuthenticationError(RuntimeError):
        status_code = 401

    async def shared_failure(*_args: Any, **_kwargs: Any) -> None:
        raise AuthenticationError("Bearer secret")

    monkeypatch.setattr(service, "_run_batch_item", shared_failure)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )
    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", lambda _record: None)

    paused = await service.retry_failed_items(record.batch_id)

    selected_id = paused.items[1].item_id
    assert paused.retry_item_ids == [selected_id]
    assert paused.status is BatchStatus.PAUSED
    called: list[str] = []

    async def complete_item(_record: BatchRecord, item: BatchItem, **_kwargs: Any) -> None:
        called.append(item.source.filename)
        item.status = BatchItemStatus.COMPLETED

    monkeypatch.setattr(service, "_run_batch_item", complete_item)
    resumed = await service.resume_batch(record.batch_id)

    assert called == ["paper-1.pdf"]
    assert resumed.retry_item_ids is None
    assert resumed.items[2].status is BatchItemStatus.QUEUED
    assert resumed.status is BatchStatus.PAUSED


@pytest.mark.asyncio
async def test_persisted_retry_scope_survives_process_restart_boundary(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(
        tmp_path,
        [BatchItemStatus.COMPLETED, BatchItemStatus.RUNNING, BatchItemStatus.QUEUED],
        batch_status=BatchStatus.RUNNING,
    )
    record.retry_item_ids = [record.items[1].item_id]
    record.current_item_id = record.items[1].item_id
    store.create(record)
    service = _service(tmp_path, store)
    called: list[str] = []

    async def complete_item(_record: BatchRecord, item: BatchItem, **_kwargs: Any) -> None:
        called.append(item.source.filename)
        item.status = BatchItemStatus.COMPLETED

    monkeypatch.setattr(service, "_run_batch_item", complete_item)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )
    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", lambda _record: None)

    resumed = await service.resume_batch(record.batch_id)

    assert called == ["paper-1.pdf"]
    assert resumed.retry_item_ids is None
    assert resumed.items[2].status is BatchItemStatus.QUEUED
    assert resumed.status is BatchStatus.PAUSED


@pytest.mark.asyncio
async def test_final_status_is_saved_only_after_csv_and_resume_retries_csv_without_model(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.QUEUED])
    store.create(record)
    service = _service(tmp_path, store)
    model_calls = 0

    async def complete_item(_record: BatchRecord, item: BatchItem, **_kwargs: Any) -> None:
        nonlocal model_calls
        model_calls += 1
        item.status = BatchItemStatus.COMPLETED

    csv_calls = 0
    csv_persisted_statuses: list[BatchStatus] = []

    def fail_final_csv(_record: BatchRecord) -> None:
        nonlocal csv_calls
        csv_calls += 1
        csv_persisted_statuses.append(store.load(record.batch_id).status)
        if csv_calls == 3:
            raise PermissionError("output is locked")

    monkeypatch.setattr(service, "_run_batch_item", complete_item)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )
    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", fail_final_csv)

    paused = await service.run_batch(record.batch_id)

    assert paused.status is BatchStatus.PAUSED
    assert store.load(record.batch_id).status is BatchStatus.PAUSED
    assert csv_persisted_statuses == [
        BatchStatus.PAUSED,
        BatchStatus.RUNNING,
        BatchStatus.RUNNING,
    ]
    assert model_calls == 1

    repaired_csv_statuses: list[BatchStatus] = []

    def repair_csv(_record: BatchRecord) -> None:
        repaired_csv_statuses.append(store.load(record.batch_id).status)

    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", repair_csv)
    completed = await service.resume_batch(record.batch_id)

    assert repaired_csv_statuses == [BatchStatus.PAUSED, BatchStatus.RUNNING]
    assert completed.status is BatchStatus.COMPLETED
    assert store.load(record.batch_id).status is BatchStatus.COMPLETED
    assert model_calls == 1


@pytest.mark.asyncio
async def test_duplicate_batch_execution_is_rejected_while_first_call_is_running(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.QUEUED])
    store.create(record)
    service = _service(tmp_path, store)
    started = asyncio.Event()
    release = asyncio.Event()

    async def blocked_item(_record: BatchRecord, item: BatchItem, **_kwargs: Any) -> None:
        started.set()
        await release.wait()
        item.status = BatchItemStatus.COMPLETED

    monkeypatch.setattr(service, "_run_batch_item", blocked_item)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )
    monkeypatch.setattr("paper_reviewer.application.service._write_batch_csv", lambda _record: None)

    first = asyncio.create_task(service.run_batch(record.batch_id))
    await started.wait()
    with pytest.raises(BatchExecutionInProgressError, match="正在执行"):
        await service.resume_batch(record.batch_id)
    release.set()

    assert (await first).status is BatchStatus.COMPLETED


@pytest.mark.asyncio
async def test_pause_cannot_overwrite_a_batch_owned_by_an_active_writer(tmp_path: Path) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.RUNNING], batch_status=BatchStatus.RUNNING)
    record.current_item_id = record.items[0].item_id
    store.create(record)
    service = _service(tmp_path, store)
    lock_store = BatchStore(service.paths.batches_dir)

    with lock_store.execution_lock(record.batch_id):
        with pytest.raises(BatchExecutionInProgressError, match="正在执行"):
            await service.pause_batch(record.batch_id)

    persisted = store.load(record.batch_id)
    assert persisted.status is BatchStatus.RUNNING
    assert persisted.items[0].status is BatchItemStatus.RUNNING


@pytest.mark.asyncio
async def test_run_batch_cancelled_error_preserves_running_item_checkpoint(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.QUEUED, BatchItemStatus.QUEUED])
    store.create(record)
    service = _service(tmp_path, store)

    async def run_item(*_args: Any, **_kwargs: Any) -> None:
        raise asyncio.CancelledError

    monkeypatch.setattr(service, "_run_batch_item", run_item)
    monkeypatch.setattr(
        "paper_reviewer.application.service.validate_source_snapshot", lambda _source: None
    )

    with pytest.raises(asyncio.CancelledError):
        await service.run_batch(record.batch_id)

    persisted = store.load(record.batch_id)
    assert persisted.status is BatchStatus.PAUSED
    assert persisted.current_item_id == persisted.items[0].item_id
    assert persisted.items[0].status is BatchItemStatus.CANCELLED
    assert persisted.items[1].status is BatchItemStatus.QUEUED


@pytest.mark.asyncio
async def test_new_batch_run_passes_frozen_source_hash_to_orchestrator(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    record = _record(tmp_path, [BatchItemStatus.RUNNING], batch_status=BatchStatus.RUNNING)
    item = record.items[0]
    item.report_path = record.request.output_dir / "existing.pdf"
    item.report_path.parent.mkdir(parents=True)
    item.report_path.write_bytes(b"pdf")
    store.create(record)
    service = _service(tmp_path, store)
    captured_hashes: list[str | None] = []

    async def start_from_snapshots(*_args: Any, **kwargs: Any) -> RunRecord:
        captured_hashes.append(kwargs.get("expected_input_hash"))
        return RunRecord(
            run_id="run-hash",
            status=RunStatus.REPORTED,
            input_path=str(item.source.path),
            input_hash=item.source.sha256,
            config_hash="b" * 64,
            rubric_id="course-paper-general-assessment@0.1.0-experimental",
            provider="openai",
            model="gpt-test",
        )

    async def load_report(_run_id: str) -> Any:
        return SimpleNamespace(
            submission_metadata=_metadata(),
            dimension_scores={},
            review=SimpleNamespace(total_score=80.0),
        )

    monkeypatch.setattr(service, "_start_review_from_snapshots", start_from_snapshots)
    monkeypatch.setattr(service, "load_report", load_report)

    await service._run_batch_item(record, item, store=store, event_sink=None)  # type: ignore[arg-type]

    assert captured_hashes == [item.source.sha256]
    assert item.status is BatchItemStatus.COMPLETED


@pytest.mark.asyncio
async def test_report_destination_is_persisted_before_publish_and_reused_after_crash(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = MemoryBatchStore()
    record = _record(
        tmp_path,
        [BatchItemStatus.RUNNING],
        batch_status=BatchStatus.RUNNING,
    )
    item = record.items[0]
    item.run_id = "run-crash"
    store.create(record)
    service = _service(tmp_path, store)
    run = RunRecord(
        run_id=item.run_id,
        status=RunStatus.REPORTED,
        input_path=str(item.source.path),
        input_hash=item.source.sha256,
        config_hash="b" * 64,
        rubric_id="course-paper-general-assessment@0.1.0-experimental",
        provider="openai",
        model="gpt-test",
    )

    async def get_run(_run_id: str) -> Any:
        return SimpleNamespace(run=run)

    async def load_report(_run_id: str) -> Any:
        return SimpleNamespace(
            submission_metadata=_metadata(),
            dimension_scores={},
            review=SimpleNamespace(total_score=80.0),
        )

    async def publish_then_crash(
        _run_id: str,
        _format: Any,
        destination: Path,
        *,
        overwrite: bool,
    ) -> Any:
        assert overwrite is False
        destination.parent.mkdir(parents=True, exist_ok=True)
        await asyncio.to_thread(
            destination.write_bytes,
            b"%PDF-1.4\napplication-owned",
        )
        raise RuntimeError("simulated process interruption after PDF publish")

    monkeypatch.setattr(service, "get_run", get_run)
    monkeypatch.setattr(service, "load_report", load_report)
    monkeypatch.setattr(service, "export_report", publish_then_crash)

    with pytest.raises(RuntimeError, match="simulated process interruption"):
        await service._run_batch_item(record, item, store=store, event_sink=None)  # type: ignore[arg-type]

    interrupted = store.load(record.batch_id)
    reserved_path = interrupted.items[0].report_path
    assert reserved_path is not None
    assert reserved_path.is_file()
    assert "__run-crash" not in reserved_path.name

    async def must_not_export_again(*_args: Any, **_kwargs: Any) -> Any:
        pytest.fail("recovery must reuse the already published reserved PDF")

    monkeypatch.setattr(service, "export_report", must_not_export_again)
    recovered_item = interrupted.items[0]
    await service._run_batch_item(  # type: ignore[arg-type]
        interrupted,
        recovered_item,
        store=store,
        event_sink=None,
    )

    persisted = store.load(record.batch_id)
    assert persisted.items[0].status is BatchItemStatus.COMPLETED
    assert persisted.items[0].report_path == reserved_path
    assert list(reserved_path.parent.glob("*课程论文评测报告.pdf")) == [reserved_path]


@pytest.mark.asyncio
async def test_metadata_correction_refreshes_local_artifacts_without_model_call(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = MemoryBatchStore()
    service, record, run_dir, old_report = _metadata_update_case(tmp_path, store)
    protected_artifacts = {
        "dimension-scores.json": b"unchanged scores",
        "reviewer-result.json": b"unchanged reviewer",
        "review-checkpoint.json": b"unchanged checkpoint",
    }
    for name, content in protected_artifacts.items():
        (run_dir / name).write_bytes(content)
    record.items[0].dimension_scores = {"task_completion": 82.0}
    record.items[0].total_score = 82.0
    store.save(record)
    calls: list[str] = []
    _install_metadata_update_fakes(monkeypatch, service, record, calls=calls)

    corrected = await service.update_submission_metadata(
        record.batch_id,
        record.items[0].item_id,
        _metadata("新姓名"),
    )

    assert corrected.items[0].metadata is not None
    assert corrected.items[0].metadata.student_name == "新姓名"
    assert calls == ["report_bundle", "pdf", "validate_pdf", "csv"]
    assert corrected.items[0].report_path is not None
    assert "新姓名" in corrected.items[0].report_path.name
    assert corrected.items[0].report_path.read_bytes() == b"new local pdf"
    assert not old_report.exists()
    assert "新姓名" in (run_dir / "submission-metadata.json").read_text(encoding="utf-8")
    assert (run_dir / "report.md").read_bytes() == b"new report.md"
    assert corrected.summary_path is not None
    assert corrected.summary_path.read_bytes() == b"new csv"
    assert corrected.workbook_path is not None
    assert corrected.workbook_path.is_file()
    workbook = load_workbook(corrected.workbook_path, data_only=False)
    try:
        worksheet = workbook["成绩汇总"]
        headers = [cell.value for cell in worksheet[1]]
        assert worksheet.cell(row=2, column=headers.index("姓名") + 1).value == "新姓名"
        assert worksheet.cell(row=2, column=headers.index("总分") + 1).value == 82
    finally:
        workbook.close()
    assert corrected.items[0].dimension_scores == {"task_completion": 82.0}
    assert corrected.items[0].total_score == 82.0
    assert {
        name: (run_dir / name).read_bytes() for name in protected_artifacts
    } == protected_artifacts


@pytest.mark.asyncio
async def test_metadata_correction_does_not_delete_unverified_output_file(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = MemoryBatchStore()
    service, record, _run_dir, old_report = _metadata_update_case(tmp_path, store)
    unrelated = old_report.with_name("教师已有文件.pdf")
    unrelated.write_bytes(b"user-owned")
    persisted = store.load(record.batch_id)
    persisted.items[0].report_path = unrelated
    store.save(persisted)
    _install_metadata_update_fakes(monkeypatch, service, persisted)

    corrected = await service.update_submission_metadata(
        persisted.batch_id,
        persisted.items[0].item_id,
        _metadata("新姓名"),
    )

    assert unrelated.read_bytes() == b"user-owned"
    assert old_report.read_bytes() == b"old local pdf"
    assert corrected.items[0].report_path not in {unrelated, old_report}


@pytest.mark.asyncio
async def test_metadata_correction_rejects_stale_dialog_snapshot(
    tmp_path: Path,
) -> None:
    store = MemoryBatchStore()
    service, record, run_dir, old_report = _metadata_update_case(tmp_path, store)
    item = record.items[0]
    assert item.metadata is not None
    expected_hash = submission_metadata_sha256(item.metadata)
    concurrent = _metadata("其他窗口修改")
    persisted = store.load(record.batch_id)
    persisted.items[0].metadata = concurrent
    store.save(persisted)
    before_artifacts = {
        path.name: path.read_bytes() for path in run_dir.iterdir() if path.is_file()
    }

    with pytest.raises(ValueError, match="其他窗口更新"):
        await service.update_submission_metadata(
            record.batch_id,
            item.item_id,
            _metadata("当前窗口修改"),
            expected_metadata_sha256=expected_hash,
        )

    assert store.load(record.batch_id).items[0].metadata == concurrent
    assert old_report.read_bytes() == b"old local pdf"
    assert {
        path.name: path.read_bytes() for path in run_dir.iterdir() if path.is_file()
    } == before_artifacts


@pytest.mark.asyncio
async def test_metadata_correction_retires_exact_schema_1_0_standard_report_name(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = MemoryBatchStore()
    service, record, run_dir, original_report = _metadata_update_case(tmp_path, store)
    original_report.unlink()
    legacy = _metadata("旧姓名")
    legacy_evidence = {
        **legacy.field_evidence,
        "paper_title": legacy.field_evidence["paper_title"].model_copy(
            update={"confidence": 0.7}
        ),
    }
    legacy = legacy.model_copy(
        update={"schema_version": "1.0", "field_evidence": legacy_evidence}
    )
    legacy_standard_name = build_report_filename(
        legacy.model_copy(update={"human_reviewed": True}),
        record.items[0].run_id or "",
    )
    legacy_report = record.request.output_dir / legacy_standard_name
    legacy_report.write_bytes(b"legacy schema 1.0 report")
    record.items[0].metadata = legacy
    record.items[0].report_path = legacy_report
    store.save(record)
    RunArtifactStore(run_dir).write_model("submission-metadata.json", legacy)
    _install_metadata_update_fakes(monkeypatch, service, record)

    corrected = await service.update_submission_metadata(
        record.batch_id,
        record.items[0].item_id,
        _metadata("新姓名"),
    )

    assert not legacy_report.exists()
    assert corrected.items[0].report_path is not None
    assert corrected.items[0].report_path.read_bytes() == b"new local pdf"


@pytest.mark.asyncio
async def test_metadata_correction_retires_allocator_collision_report_only(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = MemoryBatchStore()
    service, record, _run_dir, base_report = _metadata_update_case(tmp_path, store)
    base_report.write_bytes(b"user collision file")
    item = record.items[0]
    assert item.metadata is not None and item.run_id is not None
    collision_report = allocate_report_path(
        record.request.output_dir,
        item.metadata,
        item.run_id,
        source_filename=item.source.filename,
    )
    collision_report.write_bytes(b"batch managed collision report")
    item.report_path = collision_report
    store.save(record)
    _install_metadata_update_fakes(monkeypatch, service, record)

    await service.update_submission_metadata(
        record.batch_id,
        item.item_id,
        _metadata("新姓名"),
    )

    assert base_report.read_bytes() == b"user collision file"
    assert not collision_report.exists()


def test_metadata_transaction_never_overwrites_target_appearing_after_allocation(
    tmp_path: Path,
) -> None:
    destination = tmp_path / "new-report.pdf"
    transaction = service_module._MetadataUpdateFileTransaction()
    prepared = transaction.prepare_replacement(destination, must_be_absent=True)
    prepared.write_bytes(b"generated report")
    destination.write_bytes(b"user file created during render")

    with pytest.raises(FileExistsError, match="appeared after allocation"):
        transaction.commit()
    assert transaction.rollback() == []

    assert destination.read_bytes() == b"user file created during render"
    assert not prepared.exists()


@pytest.mark.asyncio
async def test_metadata_service_preserves_report_appearing_during_render(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = MemoryBatchStore()
    service, record, run_dir, old_report = _metadata_update_case(tmp_path, store)
    _install_metadata_update_fakes(monkeypatch, service, record)
    old_run_files = {
        path.name: path.read_bytes() for path in run_dir.iterdir() if path.is_file()
    }
    original_commit = service_module._MetadataUpdateFileTransaction.commit

    def create_user_collision_then_commit(
        transaction: service_module._MetadataUpdateFileTransaction,
    ) -> None:
        destination = next(
            operation.destination
            for operation in transaction._operations
            if operation.must_be_absent
        )
        destination.write_bytes(b"user file created while PDF rendered")
        original_commit(transaction)

    monkeypatch.setattr(
        service_module._MetadataUpdateFileTransaction,
        "commit",
        create_user_collision_then_commit,
    )

    with pytest.raises(FileExistsError, match="appeared after allocation"):
        await service.update_submission_metadata(
            record.batch_id,
            record.items[0].item_id,
            _metadata("新姓名"),
        )

    destination = record.request.output_dir / build_report_filename(
        _metadata("新姓名"),
        record.items[0].run_id or "",
        source_filename=record.items[0].source.filename,
    )
    assert destination.read_bytes() == b"user file created while PDF rendered"
    assert old_report.read_bytes() == b"old local pdf"
    assert {
        name: (run_dir / name).read_bytes() for name in old_run_files
    } == old_run_files


@pytest.mark.asyncio
async def test_metadata_thread_waits_for_worker_after_repeated_cancellation() -> None:
    started = Event()
    release = Event()

    def blocking_file_operation() -> None:
        started.set()
        assert release.wait(timeout=5)

    task = asyncio.create_task(
        service_module._metadata_update_to_thread(blocking_file_operation)
    )
    assert await asyncio.to_thread(started.wait, 2)
    task.cancel()
    await asyncio.sleep(0)
    task.cancel()
    await asyncio.sleep(0.05)
    assert not task.done()

    release.set()
    with pytest.raises(asyncio.CancelledError):
        await task


@pytest.mark.asyncio
async def test_metadata_cleanup_preserves_rollback_errors_after_repeated_cancellation() -> None:
    started = Event()
    release = Event()
    rollback_error = OSError("rollback failed")

    def blocking_rollback() -> list[BaseException]:
        started.set()
        assert release.wait(timeout=5)
        return [rollback_error]

    task = asyncio.create_task(
        service_module._metadata_cleanup_to_thread(blocking_rollback)
    )
    assert await asyncio.to_thread(started.wait, 2)
    task.cancel()
    await asyncio.sleep(0)
    task.cancel()
    await asyncio.sleep(0.05)
    assert not task.done()

    release.set()
    assert await task == [rollback_error]


@pytest.mark.asyncio
async def test_repeated_cancel_cannot_hide_incomplete_metadata_rollback(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = MemoryBatchStore()
    service, record, _run_dir, _old_report = _metadata_update_case(tmp_path, store)
    _install_metadata_update_fakes(
        monkeypatch,
        service,
        record,
        failure_stage="metadata",
    )
    rollback_started = Event()
    release_rollback = Event()

    def failing_rollback(
        _transaction: service_module._MetadataUpdateFileTransaction,
    ) -> list[BaseException]:
        rollback_started.set()
        assert release_rollback.wait(timeout=5)
        return [OSError("injected rollback failure")]

    monkeypatch.setattr(
        service_module._MetadataUpdateFileTransaction,
        "rollback",
        failing_rollback,
    )
    task = asyncio.create_task(
        service.update_submission_metadata(
            record.batch_id,
            record.items[0].item_id,
            _metadata("新姓名"),
        )
    )
    assert await asyncio.to_thread(rollback_started.wait, 2)
    task.cancel()
    await asyncio.sleep(0)
    task.cancel()
    await asyncio.sleep(0.05)
    assert not task.done()

    release_rollback.set()
    with pytest.raises(
        service_module._MetadataUpdateRollbackError,
        match="自动回滚未完全完成",
    ):
        await task


@pytest.mark.asyncio
async def test_batch_metadata_recheck_reraises_incomplete_rollback(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    record = _record(tmp_path, [BatchItemStatus.COMPLETED])
    item = record.items[0]
    item.run_id = "run-fatal-rollback"
    item.metadata = _metadata("旧姓名")
    store = MemoryBatchStore()
    store.create(record)
    service = _service(tmp_path, store)
    base_hash = submission_metadata_sha256(item.metadata)
    preview = SimpleNamespace(base_metadata_sha256=base_hash)
    prepared = SimpleNamespace(
        preview=preview,
        current=item.metadata,
        candidate=item.metadata.model_copy(update={"student_name": "新姓名"}),
    )
    monkeypatch.setattr(
        service_module,
        "_prepare_batch_metadata_recheck_item",
        lambda *_args: prepared,
    )

    async def fatal_update(*_args: Any, **_kwargs: Any) -> BatchRecord:
        raise service_module._MetadataUpdateRollbackError("rollback incomplete")

    monkeypatch.setattr(service, "_update_submission_metadata_locked", fatal_update)
    decision = MetadataRecheckDecision(
        item_id=item.item_id,
        base_metadata_sha256=base_hash,
        values={"student_name": "新姓名"},
        accepted_fields=["student_name"],
        human_reviewed=True,
    )

    with pytest.raises(
        service_module._MetadataUpdateRollbackError,
        match="rollback incomplete",
    ):
        await service.apply_batch_metadata_recheck(record.batch_id, [decision])


@pytest.mark.asyncio
async def test_batch_metadata_recheck_continues_after_ordinary_item_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    record = _record(
        tmp_path,
        [BatchItemStatus.COMPLETED, BatchItemStatus.COMPLETED],
    )
    prepared_by_id: dict[str, Any] = {}
    decisions: list[MetadataRecheckDecision] = []
    for index, item in enumerate(record.items, start=1):
        item.run_id = f"run-partial-{index}"
        item.metadata = _metadata(f"旧姓名{index}")
        base_hash = submission_metadata_sha256(item.metadata)
        prepared_by_id[item.item_id] = SimpleNamespace(
            preview=SimpleNamespace(base_metadata_sha256=base_hash),
            current=item.metadata,
            candidate=item.metadata.model_copy(update={"student_name": f"新姓名{index}"}),
        )
        decisions.append(
            MetadataRecheckDecision(
                item_id=item.item_id,
                base_metadata_sha256=base_hash,
                values={"student_name": f"新姓名{index}"},
                accepted_fields=["student_name"],
                human_reviewed=True,
            )
        )
    store = MemoryBatchStore()
    store.create(record)
    service = _service(tmp_path, store)
    monkeypatch.setattr(
        service_module,
        "_prepare_batch_metadata_recheck_item",
        lambda _runs_dir, item: prepared_by_id[item.item_id],
    )
    calls: list[str] = []

    async def update_one(
        _batch_id: str,
        item_id: str,
        metadata: SubmissionMetadata,
    ) -> BatchRecord:
        calls.append(item_id)
        if item_id == record.items[0].item_id:
            raise OSError("ordinary per-item failure")
        updated = store.load(record.batch_id)
        updated.items[1].metadata = metadata
        store.save(updated)
        return updated

    monkeypatch.setattr(service, "_update_submission_metadata_locked", update_one)

    result = await service.apply_batch_metadata_recheck(record.batch_id, decisions)

    assert calls == [item.item_id for item in record.items]
    assert result.updated_item_ids == [record.items[1].item_id]
    assert result.failed_items == {
        record.items[0].item_id: "本地重检结果应用失败；原报告和批次记录已保持不变。"
    }


@pytest.mark.asyncio
async def test_batch_metadata_recheck_rejects_missing_or_false_confirmation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    record = _record(
        tmp_path,
        [BatchItemStatus.COMPLETED, BatchItemStatus.COMPLETED],
    )
    prepared_by_id: dict[str, Any] = {}
    decisions: list[MetadataRecheckDecision] = []
    for index, item in enumerate(record.items):
        item.run_id = f"run-unconfirmed-{index}"
        item.metadata = _metadata(f"旧姓名{index}")
        base_hash = submission_metadata_sha256(item.metadata)
        prepared_by_id[item.item_id] = SimpleNamespace(
            preview=SimpleNamespace(base_metadata_sha256=base_hash),
            current=item.metadata,
            candidate=item.metadata.model_copy(update={"student_name": f"新姓名{index}"}),
        )
        payload: dict[str, Any] = {
            "item_id": item.item_id,
            "base_metadata_sha256": base_hash,
            "values": {"student_name": f"新姓名{index}"},
            "accepted_fields": ["student_name"],
        }
        if index == 1:
            payload["human_reviewed"] = False
        decisions.append(MetadataRecheckDecision.model_validate(payload))
    store = MemoryBatchStore()
    store.create(record)
    service = _service(tmp_path, store)
    monkeypatch.setattr(
        service_module,
        "_prepare_batch_metadata_recheck_item",
        lambda _runs_dir, item: prepared_by_id[item.item_id],
    )

    async def must_not_update(*_args: Any, **_kwargs: Any) -> BatchRecord:
        pytest.fail("unconfirmed metadata decision must not update files")

    monkeypatch.setattr(service, "_update_submission_metadata_locked", must_not_update)

    result = await service.apply_batch_metadata_recheck(record.batch_id, decisions)

    assert result.updated_item_ids == []
    assert result.failed_items == {
        item.item_id: "应用重新检查结果前，必须明确确认已人工核对。"
        for item in record.items
    }


@pytest.mark.asyncio
@pytest.mark.parametrize("failure_stage", ["metadata", "report_bundle", "pdf", "csv", "manifest"])
async def test_metadata_correction_rolls_back_every_published_file_on_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    failure_stage: str,
) -> None:
    store = (
        _FailAfterSaveBatchStore(tmp_path / "batches")
        if failure_stage == "manifest"
        else BatchStore(tmp_path / "batches")
    )
    service, record, run_dir, old_report = _metadata_update_case(tmp_path, store)
    old_run_files = {path.name: path.read_bytes() for path in run_dir.iterdir() if path.is_file()}
    old_pdf = old_report.read_bytes()
    assert record.summary_path is not None
    csv_path = record.summary_path
    old_csv = csv_path.read_bytes()
    old_manifest = store.load(record.batch_id)
    old_manifest_bytes = store.manifest_path(record.batch_id).read_bytes()
    _install_metadata_update_fakes(
        monkeypatch,
        service,
        record,
        failure_stage=failure_stage,
    )
    if failure_stage == "manifest":
        assert isinstance(store, _FailAfterSaveBatchStore)
        store.fail_after_next_save = True

    with pytest.raises((OSError, RuntimeError), match="injected"):
        await service.update_submission_metadata(
            record.batch_id,
            record.items[0].item_id,
            _metadata("新姓名"),
        )

    assert {name: (run_dir / name).read_bytes() for name in old_run_files} == old_run_files
    assert old_report.read_bytes() == old_pdf
    assert csv_path.read_bytes() == old_csv
    assert store.load(record.batch_id) == old_manifest
    assert store.manifest_path(record.batch_id).read_bytes() == old_manifest_bytes
    temporary_paths = await asyncio.to_thread(
        lambda: [
            path
            for path in tmp_path.rglob("*")
            if path.name.endswith((".new", ".bak"))
            or path.name.startswith(".metadata-update-build-")
        ]
    )
    assert not temporary_paths


class _FailAfterSaveBatchStore(BatchStore):
    def __init__(self, batches_root: Path) -> None:
        super().__init__(batches_root)
        self.fail_after_next_save = False

    def save(self, record: BatchRecord) -> None:
        super().save(record)
        if self.fail_after_next_save:
            self.fail_after_next_save = False
            raise OSError("injected manifest failure")


def _metadata_update_case(
    tmp_path: Path,
    store: MemoryBatchStore | BatchStore,
) -> tuple[ReviewApplicationService, BatchRecord, Path, Path]:
    record = _record(tmp_path, [BatchItemStatus.COMPLETED])
    item = record.items[0]
    item.run_id = "run-1"
    item.metadata = _metadata("旧姓名")
    output_dir = tmp_path / "reports"
    output_dir.mkdir()
    old_report = output_dir / build_report_filename(item.metadata, item.run_id)
    old_report.write_bytes(b"old local pdf")
    item.report_path = old_report
    record.summary_path = output_dir / "课程论文评测汇总_batch-test.csv"
    store.create(record)

    run_dir = tmp_path / "runs" / item.run_id
    run_dir.mkdir(parents=True)
    (run_dir / "submission-metadata.json").write_text(
        item.metadata.model_dump_json(indent=2),
        encoding="utf-8",
    )
    for name in (
        "report.json",
        "report.md",
        "evidence.json",
        "run-summary.json",
        "report-presentation.json",
    ):
        (run_dir / name).write_bytes(f"old {name}".encode())
    record.summary_path.write_bytes(b"old csv")
    return _service(tmp_path, store), record, run_dir, old_report


def _install_metadata_update_fakes(
    monkeypatch: pytest.MonkeyPatch,
    service: ReviewApplicationService,
    record: BatchRecord,
    *,
    failure_stage: str | None = None,
    calls: list[str] | None = None,
) -> None:
    observed = calls if calls is not None else []
    run = RunRecord(
        run_id="run-1",
        status=RunStatus.REPORTED,
        input_path=str(record.items[0].source.path),
        input_hash="a" * 64,
        config_hash="b" * 64,
        rubric_id="course-paper-general-assessment@0.1.0-experimental",
        provider="openai",
        model="gpt-test",
    )
    report = SimpleNamespace(
        evaluation=None,
        review=object(),
        run=run,
        rubric=record.rubric_snapshot,
        audit=object(),
        evidence=[],
        presentation_profile=None,
        dimension_scores={},
    )

    async def load_report(_run_id: str) -> Any:
        return report

    def write_bundle(*, run_dir: Path, **_kwargs: Any) -> None:
        observed.append("report_bundle")
        if failure_stage == "report_bundle":
            raise OSError("injected report_bundle failure")
        for name in ("report.json", "report.md", "evidence.json", "run-summary.json"):
            (run_dir / name).write_bytes(f"new {name}".encode())
        (run_dir / "report-presentation.json").write_bytes(b"new presentation")

    def render(markdown: str, destination: Path, *, title: str, author: str = "") -> None:
        del markdown, title, author
        observed.append("pdf")
        if failure_stage == "pdf":
            raise OSError("injected pdf failure")
        destination.write_bytes(b"new local pdf")

    def validate(destination: Path, markdown: str) -> None:
        del destination, markdown
        observed.append("validate_pdf")

    def write_csv(destination: Path, _batch: BatchRecord, _dimensions: Any) -> None:
        observed.append("csv")
        if failure_stage == "csv":
            raise OSError("injected csv failure")
        destination.write_bytes(b"new csv")

    monkeypatch.setattr(service, "load_report", load_report)
    monkeypatch.setattr("paper_reviewer.application.service.write_report_bundle", write_bundle)
    monkeypatch.setattr("paper_reviewer.application.service.render_pdf", render)
    monkeypatch.setattr("paper_reviewer.application.service.validate_pdf", validate)
    monkeypatch.setattr("paper_reviewer.application.service.write_batch_summary_csv", write_csv)
    monkeypatch.setattr(
        service,
        "_start_review_from_snapshots",
        lambda *_args, **_kwargs: pytest.fail("metadata correction must not call the model"),
    )
    if failure_stage == "metadata":
        monkeypatch.setattr(
            "paper_reviewer.application.artifacts.RunArtifactStore.write_model",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("injected metadata failure")),
        )


def _create_recheck_pdf(path: Path) -> None:
    document = pymupdf.open()
    cover = document.new_page()
    cover.insert_text((70, 80), "《人工智能导论》课程试题", fontsize=16, fontname="china-s")
    cover.insert_text((70, 130), "任课教师：李老师", fontsize=12, fontname="china-s")
    cover.insert_text((70, 165), "姓名：张三 得分：", fontsize=12, fontname="china-s")
    cover.insert_text((70, 200), "学号：202600010001", fontsize=12, fontname="china-s")
    page = document.new_page()
    page.insert_textbox(
        pymupdf.Rect(40, 65, page.rect.width - 40, 105),
        "智能时代学习能力重构：挑战、框架与实践路径",
        fontsize=18,
        fontname="china-s",
        align=pymupdf.TEXT_ALIGN_CENTER,
    )
    page.insert_text(
        (70, 145), "摘要：本文讨论大学生核心能力重构。", fontsize=11, fontname="china-s"
    )
    page.insert_textbox(
        pymupdf.Rect(70, 180, page.rect.width - 70, 700),
        "本文依据课程评价标准分析人工智能时代的学习能力、论证结构和培养路径。" * 20,
        fontsize=11,
        fontname="china-s",
    )
    document.set_metadata({"title": "示例学院"})
    document.save(path)
    document.close()


def _file_digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


@pytest.mark.asyncio
async def test_batch_metadata_recheck_preview_is_local_read_only_and_apply_is_per_item(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    request = _request(tmp_path)
    source = request.source_dir / "平台下载论文.pdf"
    _create_recheck_pdf(source)
    item = scan_batch_sources(request)[0]
    item.status = BatchItemStatus.COMPLETED
    item.run_id = "run-recheck"
    current = _metadata("张三 得分：").model_copy(
        update={
            "schema_version": "1.0",
            "paper_title": "示例学院",
            "field_evidence": {
                **_metadata().field_evidence,
                "student_name": SubmissionFieldEvidence(
                    source=SubmissionMetadataSource.COVER_LABEL,
                    confidence=0.95,
                    page=1,
                    evidence="姓名：张三 得分：",
                ),
                "paper_title": SubmissionFieldEvidence(
                    source=SubmissionMetadataSource.PDF_METADATA,
                    confidence=0.95,
                    evidence="示例学院",
                ),
            },
        }
    )
    item.metadata = current
    record = BatchRecord(
        batch_id="batch-recheck",
        status=BatchStatus.COMPLETED,
        request=request,
        rubric_snapshot=load_rubric(COURSE_RUBRIC),
        profile_snapshot=load_review_profile(COURSE_PROFILE),
        provider_snapshot=_provider(),
        items=[item],
    )
    store = BatchStore(tmp_path / "batches")
    store.create(record)
    run_dir = tmp_path / "runs" / item.run_id
    run_dir.mkdir(parents=True)
    RunArtifactStore(run_dir).write_model("submission-metadata.json", current)
    service = _service(tmp_path, store)
    monkeypatch.setattr(
        service,
        "_start_review_from_snapshots",
        lambda *_args, **_kwargs: pytest.fail("metadata recheck must not call a model"),
    )
    protected = [source, store.manifest_path(record.batch_id), run_dir / "submission-metadata.json"]
    before = {path: _file_digest(path) for path in protected}

    preview = await service.preview_batch_metadata_recheck(record.batch_id)

    assert {path: _file_digest(path) for path in protected} == before
    assert len(preview.items) == 1
    recheck_item = preview.items[0]
    by_field = {suggestion.field: suggestion for suggestion in recheck_item.suggestions}
    assert by_field["student_name"].suggested_value == "张三"
    assert by_field["paper_title"].suggested_value == (
        "智能时代学习能力重构：挑战、框架与实践路径"
    )
    assert by_field["paper_title"].evidence.source is SubmissionMetadataSource.VISIBLE_HEADING

    missing = await service.preview_batch_metadata_recheck(
        record.batch_id,
        item_ids=["missing-item"],
    )
    assert missing.items == []
    assert missing.skipped == {"missing-item": "批次中不存在该论文。"}

    captured: list[SubmissionMetadata] = []

    async def apply_local(
        _batch_id: str,
        _item_id: str,
        metadata: SubmissionMetadata,
    ) -> BatchRecord:
        captured.append(metadata.model_copy(deep=True))
        updated = store.load(record.batch_id)
        updated.items[0].metadata = metadata
        store.save(updated)
        return updated

    monkeypatch.setattr(service, "_update_submission_metadata_locked", apply_local)
    values = {
        "student_name": by_field["student_name"].suggested_value,
        "student_id": current.student_id,
        "major": current.major,
        "paper_title": by_field["paper_title"].suggested_value,
    }
    decision = MetadataRecheckDecision(
        item_id=item.item_id,
        base_metadata_sha256=recheck_item.base_metadata_sha256,
        values=values,
        accepted_fields=["student_name", "paper_title"],
        human_reviewed=True,
    )

    result = await service.apply_batch_metadata_recheck(record.batch_id, [decision])

    assert result.updated_item_ids == [item.item_id]
    assert result.failed_items == {}
    assert len(captured) == 1
    assert captured[0].student_name == "张三"
    assert captured[0].paper_title == values["paper_title"]
    assert captured[0].human_reviewed is True
    assert captured[0].field_evidence["paper_title"].source is (
        SubmissionMetadataSource.HUMAN_CORRECTION
    )
    assert submission_metadata_sha256(current) == recheck_item.base_metadata_sha256
