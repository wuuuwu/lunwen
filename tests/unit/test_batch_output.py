from __future__ import annotations

import csv
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from threading import Barrier

import pytest
from openpyxl import load_workbook

from paper_reviewer.application.batch_output import (
    BATCH_OUTPUT_OWNED_MESSAGE,
    BATCH_OUTPUT_OWNERSHIP_UNVERIFIABLE_MESSAGE,
    BATCH_OUTPUT_SUMMARY_EXISTS_MESSAGE,
    BATCH_SUMMARY_FILENAME,
    BATCH_WORKBOOK_FILENAME,
    BatchOutputConflictError,
    BatchOutputOwnedByAnotherBatchError,
    BatchOutputOwnershipUnverifiableError,
    BatchOutputSummaryExistsError,
    allocate_report_path,
    batch_output_conflict_message,
    build_report_filename,
    claim_batch_output_directory,
    is_allocated_report_filename,
    sanitize_filename_component,
    write_batch_summary_csv,
    write_batch_summary_xlsx,
)
from paper_reviewer.application.batch_store import scan_batch_sources
from paper_reviewer.config import ReviewerProfile, ReviewProfile
from paper_reviewer.domain.batch import (
    BatchItemStatus,
    BatchRecord,
    BatchReviewRequest,
)
from paper_reviewer.domain.provider import (
    ModelApiProtocol,
    ProviderSnapshot,
    endpoint_fingerprint,
)
from paper_reviewer.domain.rubric import RubricDimension, RubricProfile
from paper_reviewer.domain.submission import (
    SUBMISSION_METADATA_FIELDS,
    SubmissionFieldEvidence,
    SubmissionMetadata,
    SubmissionMetadataSource,
)


def _metadata(
    *,
    confidence: float = 0.9,
    field_confidences: dict[str, float] | None = None,
    human_reviewed: bool = False,
    **values: str,
) -> SubmissionMetadata:
    defaults = {
        "student_name": "张三",
        "student_id": "20260001",
        "paper_title": "课程论文题目",
    }
    defaults.update(values)
    return SubmissionMetadata(
        **defaults,
        field_evidence={
            field: SubmissionFieldEvidence(
                source=SubmissionMetadataSource.COVER_LABEL,
                confidence=(field_confidences or {}).get(field, confidence),
                page=1,
            )
            for field in SUBMISSION_METADATA_FIELDS
        },
        human_reviewed=human_reviewed,
    )


def _record(tmp_path: Path, metadata: SubmissionMetadata) -> BatchRecord:
    source_dir = tmp_path / "papers"
    source_dir.mkdir()
    (source_dir / "=paper.pdf").write_bytes(b"paper")
    request = BatchReviewRequest(
        source_dir=source_dir,
        output_dir=tmp_path / "output",
        provider="openai",
        model="gpt-test",
        rubric=Path("rubric.yaml"),
        profile=Path("profile.yaml"),
    )
    item = scan_batch_sources(request)[0]
    item.status = BatchItemStatus.COMPLETED
    item.run_id = "12345678abcdef"
    item.metadata = metadata
    item.dimension_scores = {"completion": 88, "writing": 77}
    item.total_score = 84.5
    item.grade = "良好"
    item.conclusion = "达到课程要求"
    item.report_path = request.output_dir / "output.pdf"
    item.error = "=danger"
    base_url = "https://api.openai.com/v1"
    return BatchRecord(
        batch_id="batch-1",
        request=request,
        rubric_snapshot=RubricProfile(
            rubric_id="course-test",
            version="1",
            title="课程论文",
        ),
        profile_snapshot=ReviewProfile(
            profile_id="course-test",
            version="1",
            reviewers=[
                ReviewerProfile(
                    reviewer_id="reviewer",
                    title="评阅人",
                    description="测试",
                )
            ],
        ),
        provider_snapshot=ProviderSnapshot(
            provider_ref="openai",
            display_name="OpenAI",
            protocol=ModelApiProtocol.CHAT_COMPLETIONS,
            base_url=base_url,
            endpoint_fingerprint=endpoint_fingerprint(
                base_url,
                ModelApiProtocol.CHAT_COMPLETIONS,
            ),
            model=request.model,
        ),
        items=[item],
    )


def _workbook_dimensions() -> list[RubricDimension]:
    return [
        RubricDimension(
            dimension_id="completion",
            title="共同维度",
            description="课程任务完成情况",
            weight=60,
            minimum_score=0,
            maximum_score=100,
            checks=["检查任务完成情况"],
        ),
        RubricDimension(
            dimension_id="writing",
            title="共同维度",
            description="文字表达情况",
            weight=40,
            minimum_score=10,
            maximum_score=90,
            checks=["检查文字表达"],
        ),
    ]


def test_report_filename_normalizes_windows_names_and_limits_utf16() -> None:
    metadata = _metadata(
        student_name="  张：三  ",
        student_id="CON",
        paper_title="😀" * 200 + "?",
    )

    filename = build_report_filename(metadata, "abcdef123456")

    assert filename.startswith("张_三__CON_")
    assert filename.endswith("_课程论文评测报告.pdf")
    assert len(filename.encode("utf-16-le")) // 2 <= 240
    assert not any(character in filename for character in '<>:"/\\|?*')


def test_low_confidence_critical_metadata_uses_safe_pending_name() -> None:
    metadata = _metadata(field_confidences={"student_name": 0.5})

    filename = build_report_filename(
        metadata,
        "abcdef123456",
        source_filename="原始交稿.pdf",
    )

    assert filename == "待核对论文__待核对__abcdef12_课程论文评测报告.pdf"
    assert "原始交稿" not in filename
    assert "张三" not in filename
    assert "20260001" not in filename


def test_human_review_keeps_standard_filename() -> None:
    reviewed = _metadata(
        field_confidences={"student_name": 0.2, "paper_title": 0.2},
        human_reviewed=True,
    )
    assert build_report_filename(reviewed, "abcdef123456").startswith(
        "张三_20260001_课程论文题目"
    )


def test_sanitize_filename_component_uses_nfkc_and_placeholder() -> None:
    assert sanitize_filename_component(" \uff21\uff22\uff23 ", fallback="missing") == "ABC"
    assert sanitize_filename_component("...", fallback="missing") == "missing"


def test_allocate_report_path_never_overwrites_and_uses_run_suffix(tmp_path: Path) -> None:
    output = tmp_path / "output"
    output.mkdir()
    metadata = _metadata()
    original = output / build_report_filename(metadata, "12345678abcdef")
    original.write_bytes(b"keep")

    allocated = allocate_report_path(output, metadata, "12345678abcdef")
    allocated.write_bytes(b"second")
    third = allocate_report_path(output, metadata, "12345678abcdef")

    assert allocated.name.endswith("__12345678_课程论文评测报告.pdf")
    assert third.name.endswith("__12345678_2_课程论文评测报告.pdf")
    assert original.read_bytes() == b"keep"


def test_allocated_report_name_recognizer_accepts_only_current_run_suffixes(
    tmp_path: Path,
) -> None:
    output = tmp_path / "reports"
    output.mkdir()
    metadata = _metadata()
    run_id = "12345678abcdef"
    base = build_report_filename(metadata, run_id)
    stem = base.removesuffix("_课程论文评测报告.pdf")

    assert is_allocated_report_filename(output, base, metadata, run_id)
    assert is_allocated_report_filename(
        output,
        f"{stem}__12345678_课程论文评测报告.pdf",
        metadata,
        run_id,
    )
    assert is_allocated_report_filename(
        output,
        f"{stem}__12345678_27_课程论文评测报告.pdf",
        metadata,
        run_id,
    )
    assert not is_allocated_report_filename(
        output,
        f"{stem}__87654321_课程论文评测报告.pdf",
        metadata,
        run_id,
    )
    assert not is_allocated_report_filename(
        output,
        f"{stem}__12345678_1_课程论文评测报告.pdf",
        metadata,
        run_id,
    )
    assert not is_allocated_report_filename(
        output,
        "用户文件__12345678_2_课程论文评测报告.pdf",
        metadata,
        run_id,
    )


def test_batch_csv_is_bom_atomic_dynamic_and_formula_safe(tmp_path: Path) -> None:
    metadata = _metadata(student_name="=HYPERLINK(\"x\")", paper_title="+SUM(1,1)")
    record = _record(tmp_path, metadata)
    destination = tmp_path / "output" / "summary.csv"

    write_batch_summary_csv(
        destination,
        record,
        [("completion", "课程任务完成度"), ("writing", "文字表达")],
    )

    content = destination.read_bytes()
    assert content.startswith(b"\xef\xbb\xbf")
    with destination.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["姓名"].startswith("'=HYPERLINK")
    assert rows[0]["题目"].startswith("'+SUM")
    assert rows[0]["原文件名"] == "'=paper.pdf"
    assert rows[0]["错误摘要"] == "'=danger"
    assert rows[0]["课程任务完成度"] == "88"
    assert rows[0]["文字表达"] == "77"
    assert rows[0]["元数据置信度"] == "0.90"
    assert rows[0]["待核对字段"] == ""
    assert rows[0]["人工已核对"] == "否"
    assert rows[0]["重复PDF内容"] == "否"
    assert not destination.with_suffix(".csv.tmp").exists()


def test_batch_csv_lists_pending_fields_and_human_confirmation(tmp_path: Path) -> None:
    (tmp_path / "pending").mkdir()
    (tmp_path / "confirmed").mkdir()
    pending = _record(
        tmp_path / "pending",
        _metadata(field_confidences={"student_id": 0.4, "paper_title": 0.3}),
    )
    confirmed = _record(
        tmp_path / "confirmed",
        _metadata(
            field_confidences={"student_id": 0.4, "paper_title": 0.3},
            human_reviewed=True,
        ),
    )
    pending_destination = tmp_path / "pending-output" / "summary.csv"
    confirmed_destination = tmp_path / "confirmed-output" / "summary.csv"

    write_batch_summary_csv(pending_destination, pending, [])
    write_batch_summary_csv(confirmed_destination, confirmed, [])

    with pending_destination.open(encoding="utf-8-sig", newline="") as handle:
        pending_row = next(csv.DictReader(handle))
    with confirmed_destination.open(encoding="utf-8-sig", newline="") as handle:
        confirmed_row = next(csv.DictReader(handle))
    assert pending_row["元数据待核对"] == "是"
    assert pending_row["待核对字段"] == "学号、题目"
    assert pending_row["人工已核对"] == "否"
    assert confirmed_row["元数据待核对"] == "否"
    assert confirmed_row["待核对字段"] == ""
    assert confirmed_row["人工已核对"] == "是"


def test_batch_xlsx_is_styled_dynamic_typed_and_formula_safe(tmp_path: Path) -> None:
    metadata = _metadata(
        student_name="=HYPERLINK(\"https://invalid.example\")",
        student_id="0012300",
        paper_title="+SUM(1,1)",
        field_confidences={"paper_title": 0.4},
    )
    record = _record(tmp_path, metadata)
    failed = record.items[0].model_copy(deep=True)
    failed.item_id = "failed-item"
    failed.status = BatchItemStatus.FAILED
    failed.metadata = None
    failed.dimension_scores = {}
    failed.total_score = None
    failed.grade = None
    failed.conclusion = None
    failed.report_path = None
    failed.error = "@external-error"
    record.items.append(failed)
    destination = record.request.output_dir / BATCH_WORKBOOK_FILENAME

    write_batch_summary_xlsx(destination, record, _workbook_dimensions())

    workbook = load_workbook(destination, data_only=False, keep_links=True)
    try:
        assert workbook.sheetnames == ["成绩汇总"]
        worksheet = workbook["成绩汇总"]
        headers = [cell.value for cell in worksheet[1]]
        assert "共同维度（completion）" in headers
        assert "共同维度（writing）" in headers
        assert "专业" not in headers
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref == worksheet.tables["CoursePaperScores"].ref
        assert worksheet.tables["CoursePaperScores"].tableStyleInfo.showRowStripes
        assert max(dimension.width or 0 for dimension in worksheet.column_dimensions.values()) <= 48

        by_header = {value: index for index, value in enumerate(headers, start=1)}
        student_id = worksheet.cell(row=2, column=by_header["学号"])
        confidence = worksheet.cell(row=2, column=by_header["元数据置信度"])
        first_score = worksheet.cell(row=2, column=by_header["共同维度（completion）"])
        total = worksheet.cell(row=2, column=by_header["总分"])
        assert student_id.value == "0012300"
        assert student_id.data_type == "s"
        assert confidence.value == pytest.approx(0.7333333333)
        assert confidence.data_type == "n"
        assert first_score.value == 88
        assert first_score.data_type == "n"
        assert total.value == 84.5
        assert total.data_type == "n"
        assert worksheet.cell(row=3, column=by_header["总分"]).value is None
        assert worksheet.cell(row=3, column=by_header["任务状态"]).value == "failed"

        completion_header = worksheet.cell(
            row=1,
            column=by_header["共同维度（completion）"],
        )
        writing_header = worksheet.cell(
            row=1,
            column=by_header["共同维度（writing）"],
        )
        assert completion_header.comment is not None
        assert "权重：60%" in completion_header.comment.text
        assert "分值范围：0–100" in completion_header.comment.text
        assert writing_header.comment is not None
        assert "分值范围：10–90" in writing_header.comment.text

        assert not workbook._external_links
        for row in worksheet.iter_rows():
            for cell in row:
                assert cell.data_type != "f"
                assert cell.hyperlink is None
        assert worksheet.cell(row=2, column=by_header["姓名"]).value.startswith("=")
        assert worksheet.cell(row=2, column=by_header["原文件名"]).value == "=paper.pdf"
        assert worksheet.cell(row=2, column=by_header["题目"]).value.startswith("+")
        assert worksheet.cell(row=3, column=by_header["错误摘要"]).value.startswith("@")
    finally:
        workbook.close()


def test_batch_xlsx_replace_failure_preserves_old_file_and_cleans_temp(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    record = _record(tmp_path, _metadata())
    output_dir = record.request.output_dir
    output_dir.mkdir(parents=True)
    claim_batch_output_directory(output_dir, record.batch_id)
    destination = output_dir / BATCH_WORKBOOK_FILENAME
    destination.write_bytes(b"old workbook")

    def deny_replace(_source: Path, _destination: Path) -> None:
        raise PermissionError("locked by spreadsheet application")

    monkeypatch.setattr("paper_reviewer.application.batch_output.os.replace", deny_replace)

    with pytest.raises(PermissionError):
        write_batch_summary_xlsx(destination, record, _workbook_dimensions())

    assert destination.read_bytes() == b"old workbook"
    assert not list(output_dir.glob(f".{BATCH_WORKBOOK_FILENAME}.*.tmp.xlsx"))


def test_batch_csv_does_not_replace_a_summary_claimed_by_another_batch(
    tmp_path: Path,
) -> None:
    destination = tmp_path / "output" / "summary.csv"
    (tmp_path / "first").mkdir()
    (tmp_path / "second").mkdir()
    first = _record(tmp_path / "first", _metadata())
    second = _record(tmp_path / "second", _metadata())
    second.batch_id = "batch-2"
    write_batch_summary_csv(destination, first, [])
    original = destination.read_bytes()

    with pytest.raises(FileExistsError, match="另一个课程论文批次"):
        write_batch_summary_csv(destination, second, [])

    assert destination.read_bytes() == original


def test_batch_csv_dimension_headers_are_formula_safe(tmp_path: Path) -> None:
    record = _record(tmp_path, _metadata())
    destination = tmp_path / "output" / "summary.csv"

    write_batch_summary_csv(destination, record, [("unsafe", "=HYPERLINK(\"x\")")])

    with destination.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        assert "'=HYPERLINK(\"x\")" in (reader.fieldnames or [])


def test_batch_csv_refuses_to_overwrite_another_batch_output(tmp_path: Path) -> None:
    first = _record(tmp_path, _metadata())
    destination = tmp_path / "output" / "课程论文评测汇总.csv"
    write_batch_summary_csv(destination, first, [])
    original = destination.read_bytes()
    second = first.model_copy(update={"batch_id": "batch-2"}, deep=True)

    with pytest.raises(FileExistsError, match="另一个课程论文批次"):
        write_batch_summary_csv(destination, second, [])

    assert destination.read_bytes() == original


def test_batch_output_conflict_probe_is_static_and_has_no_side_effects(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"

    assert batch_output_conflict_message(output_dir) is None
    assert not output_dir.exists()

    output_dir.mkdir()
    summary = output_dir / BATCH_SUMMARY_FILENAME
    summary.write_bytes(b"existing")
    before = tuple(output_dir.iterdir())

    assert batch_output_conflict_message(output_dir) == BATCH_OUTPUT_SUMMARY_EXISTS_MESSAGE
    assert tuple(output_dir.iterdir()) == before
    with pytest.raises(BatchOutputSummaryExistsError) as captured:
        claim_batch_output_directory(output_dir, "new-batch")

    assert str(captured.value) == BATCH_OUTPUT_SUMMARY_EXISTS_MESSAGE
    assert tuple(output_dir.iterdir()) == before


def test_batch_output_conflict_probe_also_protects_unowned_xlsx(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    workbook = output_dir / BATCH_WORKBOOK_FILENAME
    workbook.write_bytes(b"existing workbook")
    before = tuple(output_dir.iterdir())

    assert batch_output_conflict_message(output_dir) == BATCH_OUTPUT_SUMMARY_EXISTS_MESSAGE
    with pytest.raises(BatchOutputSummaryExistsError):
        claim_batch_output_directory(output_dir, "new-batch")

    assert tuple(output_dir.iterdir()) == before
    assert workbook.read_bytes() == b"existing workbook"


def test_batch_output_claim_allows_only_its_batch_to_resume(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    output_dir.mkdir()

    claim_batch_output_directory(output_dir, "batch-one")

    assert batch_output_conflict_message(output_dir) == BATCH_OUTPUT_OWNED_MESSAGE
    assert batch_output_conflict_message(output_dir, batch_id="batch-one") is None
    assert (
        batch_output_conflict_message(output_dir, batch_id="batch-two")
        == BATCH_OUTPUT_OWNED_MESSAGE
    )
    claim_batch_output_directory(output_dir, "batch-one")
    with pytest.raises(BatchOutputOwnedByAnotherBatchError) as captured:
        claim_batch_output_directory(output_dir, "batch-two")

    assert str(captured.value) == BATCH_OUTPUT_OWNED_MESSAGE


def test_competing_output_claims_have_one_winner_and_one_safe_conflict(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    barrier = Barrier(2)

    def compete(batch_id: str) -> BaseException | None:
        barrier.wait()
        try:
            claim_batch_output_directory(output_dir, batch_id)
        except BaseException as error:
            return error
        return None

    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(compete, ("batch-one", "batch-two")))

    assert sum(result is None for result in results) == 1
    conflict = next(result for result in results if result is not None)
    assert isinstance(conflict, BatchOutputConflictError)
    assert str(conflict) in {
        BATCH_OUTPUT_OWNED_MESSAGE,
        "该报告输出目录的批次归属标记无法验证；请选择新的空目录。",
    }
    assert "output" not in str(conflict).casefold()


def test_corrupt_owner_is_a_static_explicit_conflict(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    owner = output_dir / f".{BATCH_SUMMARY_FILENAME}.owner"
    owner.write_bytes(b"\xffprivate-path")

    assert (
        batch_output_conflict_message(output_dir, batch_id="existing-batch")
        == BATCH_OUTPUT_OWNERSHIP_UNVERIFIABLE_MESSAGE
    )
    with pytest.raises(BatchOutputOwnershipUnverifiableError) as captured:
        claim_batch_output_directory(output_dir, "existing-batch")

    assert str(captured.value) == BATCH_OUTPUT_OWNERSHIP_UNVERIFIABLE_MESSAGE
    assert "private-path" not in str(captured.value)
